"""
Prognose 2026 — ODS + Privé
============================
Berekent:
  1. ODS winst na belasting (€70k omzet excl. BTW, eenmanszaak)
  2. Maximale ZZP opname + maandaanbeveling voor Mei-Dec
  3. ABN privé kosten: werkelijk Jan-Apr + projectie Mei-Dec
  4. Spaarprognose

Run:
    cd business
    python -X utf8 prognose_2026.py
"""

import sys
sys.stdout.reconfigure(encoding="utf-8")

import sqlite3
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from config import OUTPUT_DIR

YEAR       = "2026"
DB_PATH    = Path(__file__).parent.parent / "bookkeeping" / "finance.db"
OMZET_ODS  = 70_000.00
BTW_PCT    = 0.21

# ── 2026 NL fiscale parameters (eenmanszaak, schatting) ──────────────────────
ZELFSTANDIGENAFTREK  = 2_470.00   # 2026 (gelijk aan 2025)
MKB_VRIJSTELLING_PCT = 0.127       # 12.7%
BOX1_GRENS1          = 38_441.00
BOX1_T1              = 0.3582
BOX1_GRENS2          = 76_817.00
BOX1_T2              = 0.3748
BOX1_T3              = 0.4950
AHK_MAX              = 3_473.00
AHK_AFBOUW_START     = 24_813.00
AHK_AFBOUW_PCT       = 0.06337
ARK_MAX              = 5_158.00
ARK_OPBOUW_GRENS     = 24_820.00
ARK_PLATEAU_GRENS    = 43_071.00
ARK_AFBOUW_PCT       = 0.0651
ARK_AFBOUW_MAX       = 124_935.00

EURO = "€ #,##0;-€ #,##0"
EURO2 = "€ #,##0.00;-€ #,##0.00"

MAANDEN_NL = {
    1:"Jan", 2:"Feb", 3:"Mrt", 4:"Apr", 5:"Mei", 6:"Jun",
    7:"Jul", 8:"Aug", 9:"Sep", 10:"Okt", 11:"Nov", 12:"Dec",
}

C_DARK    = "1F4E79"
C_MED     = "2E75B6"
C_SUB     = "BDD7EE"
C_GREEN   = "375623"
C_RED     = "C00000"
C_ORANGE  = "ED7D31"
C_ALT     = "F2F2F2"
C_ACTUAL  = "DEEAF1"  # blauwtint voor werkelijke maanden
C_PROJ    = "FFF2CC"  # geeltint voor prognose maanden
WHITE     = "FFFFFF"

INCOME_CATS = {
    "Salaris", "ZZP Opname", "ZZP Inkomen", "Zorgtoeslag",
    "DUO / Studiefinanciering", "Familie & Giften", "Inkomsten Overig",
}


def _fill(hex_c):
    return PatternFill("solid", fgColor=hex_c)


def _font(bold=False, color="000000", size=10):
    return Font(bold=bold, color=color, name="Arial", size=size)


# ── Fiscale berekeningen ──────────────────────────────────────────────────────

def _box1_bruto(bw):
    if bw <= BOX1_GRENS1:
        return bw * BOX1_T1
    elif bw <= BOX1_GRENS2:
        return BOX1_GRENS1 * BOX1_T1 + (bw - BOX1_GRENS1) * BOX1_T2
    else:
        return (BOX1_GRENS1 * BOX1_T1
                + (BOX1_GRENS2 - BOX1_GRENS1) * BOX1_T2
                + (bw - BOX1_GRENS2) * BOX1_T3)


def _ahk(bw):
    if bw <= AHK_AFBOUW_START:
        return AHK_MAX
    afbouw = AHK_AFBOUW_PCT * (bw - AHK_AFBOUW_START)
    return max(0.0, AHK_MAX - afbouw)


def _ark(bw):
    if bw <= 11_491:
        return bw * 0.08231
    elif bw <= ARK_OPBOUW_GRENS:
        return 945 + (bw - 11_491) / (ARK_OPBOUW_GRENS - 11_491) * (ARK_MAX - 945)
    elif bw <= ARK_PLATEAU_GRENS:
        return ARK_MAX
    elif bw <= ARK_AFBOUW_MAX:
        return max(0.0, ARK_MAX - ARK_AFBOUW_PCT * (bw - ARK_PLATEAU_GRENS))
    return 0.0


def bereken_ods(omzet, kosten_ods):
    winst   = omzet - kosten_ods
    winst_na_za = max(0.0, winst - ZELFSTANDIGENAFTREK)
    mkb     = winst_na_za * MKB_VRIJSTELLING_PCT
    bw      = winst_na_za - mkb

    box1    = _box1_bruto(bw)
    ahk     = _ahk(bw)
    ark     = _ark(bw)
    ib      = max(0.0, box1 - ahk - ark)

    return {
        "omzet":             omzet,
        "btw_reservering":   omzet * BTW_PCT,
        "kosten_ods":        kosten_ods,
        "winst":             winst,
        "zelfstandigenaftrek": ZELFSTANDIGENAFTREK,
        "winst_na_za":       winst_na_za,
        "mkb_vrijstelling":  mkb,
        "belastbare_winst":  bw,
        "box1_bruto":        box1,
        "ahk":               ahk,
        "ark":               ark,
        "ib_te_betalen":     ib,
        "beschikbaar":       winst - ib,
    }


# ── DB queries ────────────────────────────────────────────────────────────────

def _query(sql, params=()):
    with sqlite3.connect(DB_PATH) as c:
        return pd.read_sql_query(sql, c, params=params)


def load_abn_2026():
    """Maand × categorie totalen voor ABN + Revolut, 2026."""
    return _query("""
        SELECT month, category, SUM(amount) AS total
        FROM transactions
        WHERE year='2026' AND category != 'Interne Overboeking'
          AND source IN ('ABN','Revolut')
        GROUP BY month, category
    """)


def load_knab_2026():
    """Knab 2026 kosten (negatieve transacties)."""
    return _query("""
        SELECT month, SUM(amount) AS total
        FROM transactions
        WHERE year='2026' AND source='Knab'
          AND category NOT IN ('Interne Overboeking','ZZP Inkomen','Inkomsten Overig')
          AND amount < 0
        GROUP BY month
    """)


# ── Excel helpers ─────────────────────────────────────────────────────────────

def _hdr(cell, text, bg=C_DARK, fg=WHITE, bold=True, size=10, align="left"):
    cell.value = text
    cell.font  = Font(bold=bold, color=fg, name="Arial", size=size)
    cell.fill  = _fill(bg)
    cell.alignment = Alignment(horizontal=align, vertical="center")


def _num(cell, value, fmt=EURO, align="right", bg=None):
    cell.value        = value
    cell.number_format = fmt
    cell.alignment    = Alignment(horizontal=align)
    cell.font         = _font()
    if bg:
        cell.fill = _fill(bg)


def _lbl(cell, text, bold=False, bg=None, indent=0, color="000000"):
    cell.value     = ("  " * indent) + text
    cell.font      = _font(bold=bold, color=color)
    cell.alignment = Alignment(horizontal="left")
    if bg:
        cell.fill = _fill(bg)


# ── Sheet 1: ODS Berekening ───────────────────────────────────────────────────

def write_ods_sheet(ws, ods, knab_werkelijk, knab_prognose, zzp_opname_reeds):
    ws.column_dimensions["A"].width = 44
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 32
    ws.freeze_panes = "A3"

    r = 1

    def title(text):
        nonlocal r
        ws.merge_cells(f"A{r}:C{r}")
        c = ws.cell(r, 1, text)
        c.font = Font(bold=True, color=WHITE, name="Arial", size=13)
        c.fill = _fill(C_DARK)
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[r].height = 26
        r += 1

    def section(text):
        nonlocal r
        for col in (1, 2, 3):
            ws.cell(r, col).fill = _fill(C_MED)
        c = ws.cell(r, 1, text)
        c.font = _font(bold=True, color=WHITE)
        ws.row_dimensions[r].height = 18
        r += 1

    def row_item(label, amount, note="", alt=False, bold=False, color="000000"):
        nonlocal r
        bg = C_ALT if alt else None
        lc = ws.cell(r, 1)
        ac = ws.cell(r, 2)
        nc = ws.cell(r, 3)
        _lbl(lc, label, bold=bold, bg=bg, indent=1, color=color)
        _num(ac, amount, bg=bg)
        if note:
            nc.value = note
            nc.font  = _font(color="595959", size=9)
            nc.alignment = Alignment(horizontal="left")
            if bg:
                nc.fill = _fill(bg)
        r += 1
        return amount

    def row_subtotal(label, amount, bg=C_SUB, fg="000000", size=10):
        nonlocal r
        lc = ws.cell(r, 1)
        ac = ws.cell(r, 2)
        lc.value = label
        lc.font  = Font(bold=True, color=fg, name="Arial", size=size)
        lc.fill  = _fill(bg)
        _num(ac, amount, bg=bg)
        ac.font  = Font(bold=True, color=fg, name="Arial", size=size)
        if size > 10:
            ws.row_dimensions[r].height = size + 8
        r += 1
        return amount

    def gap():
        nonlocal r
        r += 1

    title(f"ODS Winstberekening 2026  —  gebaseerd op €{OMZET_ODS:,.0f} omzet excl. BTW")
    gap()

    # Omzet
    section("OMZET")
    row_item("Omzet excl. BTW  (opgegeven)", ods["omzet"], alt=True, bold=True)
    row_item("BTW 21%  (reserveren, niet winst)", ods["btw_reservering"],
             note="Apart reserveren op spaarrekening", alt=False)
    row_item("BTW-plichtige omzet (inclusief BTW)", ods["omzet"] + ods["btw_reservering"],
             note="Wat je factureerde aan klanten")
    gap()

    # Kosten
    section("BEDRIJFSKOSTEN (Knab)")
    row_item("Werkelijk Jan-Apr  (uit boekhouding)", -knab_werkelijk,
             note="Zakelijke kosten excl. BTW", alt=True)
    row_item("Prognose Mei-Dec  (geprojecteerd)", -knab_prognose,
             note=f"Avg. {knab_werkelijk/4:.0f}/mnd × 8")
    row_subtotal("Totaal bedrijfskosten", -ods["kosten_ods"])
    gap()

    # Winst
    section("WINST EN FISCALE AFTREKPOSTEN")
    row_item("Winst uit onderneming", ods["winst"], bold=True, alt=True)
    row_item("Zelfstandigenaftrek  (2026)", -ods["zelfstandigenaftrek"],
             note="Vereist: >1225 uur ondernemen")
    row_item("Winst na zelfstandigenaftrek", ods["winst_na_za"])
    row_item("MKB-winstvrijstelling  (12,7%)", -ods["mkb_vrijstelling"],
             note="Vrijgesteld van belasting", alt=True)
    row_subtotal("Belastbare winst  (Box 1)", ods["belastbare_winst"])
    gap()

    # IB
    section("INKOMSTENBELASTING 2026  (schatting, eenmanszaak)")
    box1_s1 = min(ods["belastbare_winst"], BOX1_GRENS1) * BOX1_T1
    box1_s2 = max(0.0, min(ods["belastbare_winst"], BOX1_GRENS2) - BOX1_GRENS1) * BOX1_T2
    row_item(f"Schijf 1  (≤ €{BOX1_GRENS1:,.0f} × {BOX1_T1*100:.2f}%)", box1_s1, alt=True)
    if box1_s2 > 0:
        row_item(f"Schijf 2  (> €{BOX1_GRENS1:,.0f} × {BOX1_T2*100:.2f}%)", box1_s2)
    row_item("Box 1 bruto belasting", ods["box1_bruto"])
    row_item("Algemene heffingskorting", -ods["ahk"],
             note="Inkomensafhankelijk", alt=True)
    row_item("Arbeidskorting", -ods["ark"],
             note="Inkomensafhankelijk")
    row_subtotal("IB/PVV te betalen  (schatting)", ods["ib_te_betalen"], bg="FFD6E0")
    row_item("Maandelijkse IB-reservering", -(ods["ib_te_betalen"] / 12),
             note="Apart zetten op spaarrekening", alt=True)
    row_item("BTW-reservering per kwartaal", -(ods["btw_reservering"] / 4),
             note="Elk kwartaal apart zetten")
    gap()

    # Beschikbaar
    section("BESCHIKBAAR VOOR PRIVÉ")
    row_item("Winst uit onderneming", ods["winst"], alt=True)
    row_item("IB te betalen", -ods["ib_te_betalen"])
    row_subtotal("Beschikbaar voor privé  (heel 2026)", ods["beschikbaar"],
                 bg=C_GREEN, fg=WHITE, size=11)
    gap()
    row_item("Reeds opgenomen  (ZZP Opname Jan-Apr)", -zzp_opname_reeds,
             note="Werkelijk uit boekhouding", alt=True)
    restant = ods["beschikbaar"] - zzp_opname_reeds
    row_subtotal("Nog op te nemen  (Mei-Dec)", restant, bg=C_SUB)
    maand_opname = restant / 8
    row_item("→ Aanbevolen opname per maand  (8 mnd)", maand_opname,
             note="Mei t/m Dec", bold=True)


# ── Sheet 2: ABN Kosten Prognose ─────────────────────────────────────────────

def write_abn_sheet(ws, abn_df, maand_opname_aanbevolen, salaris_maand, zorgtoeslag_maand):
    ACTUAL_MONTHS = [f"2026-{m:02d}" for m in range(1, 5)]
    PROJ_MONTHS   = [f"2026-{m:02d}" for m in range(5, 13)]
    ALL_MONTHS    = ACTUAL_MONTHS + PROJ_MONTHS

    # Pivot: category × month
    pivot = abn_df.pivot_table(
        index="category", columns="month", values="total", aggfunc="sum", fill_value=0
    )

    # Separate income and costs
    all_cats = list(pivot.index)
    inc_cats  = [c for c in all_cats if c in INCOME_CATS]
    cost_cats = sorted([c for c in all_cats if c not in INCOME_CATS],
                       key=lambda c: pivot.loc[c, ACTUAL_MONTHS].sum())

    # Monthly averages for projection (based on Jan-Apr actual)
    def avg_monthly(cat):
        vals = [pivot.loc[cat, m] if m in pivot.columns else 0.0 for m in ACTUAL_MONTHS]
        return sum(vals) / len(ACTUAL_MONTHS)

    n_months = len(ALL_MONTHS)
    col_offset = 2  # A=label, B..onwards=months

    # Column widths
    ws.column_dimensions["A"].width = 26
    for i in range(1, n_months + 3):
        ws.column_dimensions[get_column_letter(col_offset + i - 1)].width = 9
    ws.column_dimensions[get_column_letter(col_offset + n_months)].width = 12  # Jaar
    ws.freeze_panes = "B3"

    r = 1

    # Title row
    ws.merge_cells(f"A{r}:{get_column_letter(col_offset + n_months)}{r}")
    tc = ws.cell(r, 1, f"ABN Privé Kosten Prognose {YEAR}  —  Werkelijk Jan-Apr | Prognose Mei-Dec")
    tc.font  = Font(bold=True, color=WHITE, name="Arial", size=12)
    tc.fill  = _fill(C_DARK)
    tc.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[r].height = 24
    r += 1

    # Month header row
    ws.cell(r, 1, "Categorie").font = _font(bold=True, color=WHITE)
    ws.cell(r, 1).fill = _fill(C_DARK)
    for i, m in enumerate(ALL_MONTHS):
        mn = int(m.split("-")[1])
        c = ws.cell(r, col_offset + i, MAANDEN_NL[mn])
        c.font      = _font(bold=True, color=WHITE)
        c.fill      = _fill(C_ACTUAL if m in ACTUAL_MONTHS else C_DARK)
        c.alignment = Alignment(horizontal="center")
    # Jaar header
    jc = ws.cell(r, col_offset + n_months, "Jaar")
    jc.font  = _font(bold=True, color=WHITE)
    jc.fill  = _fill(C_DARK)
    jc.alignment = Alignment(horizontal="center")
    ws.row_dimensions[r].height = 18
    r += 1

    def write_row(label, vals_by_month, projected_months, bold=False, bg=None, color="000000"):
        lc = ws.cell(r, 1, label)
        lc.font      = _font(bold=bold, color=color)
        lc.alignment = Alignment(horizontal="left")
        if bg:
            lc.fill = _fill(bg)
        jaar = 0
        for i, m in enumerate(ALL_MONTHS):
            v = vals_by_month.get(m, 0.0)
            cc = ws.cell(r, col_offset + i, v)
            cc.number_format = EURO
            cc.alignment     = Alignment(horizontal="right")
            cc.font          = _font(bold=bold, color=color)
            if m in projected_months:
                cc.fill = _fill(C_PROJ)
            elif bg:
                cc.fill = _fill(bg)
            jaar += v
        # Jaar totaal
        yc = ws.cell(r, col_offset + n_months, jaar)
        yc.number_format = EURO
        yc.alignment     = Alignment(horizontal="right")
        yc.font          = _font(bold=True, color=color)
        if bg:
            yc.fill = _fill(bg)
        return jaar

    # ── INKOMSTEN ─────────────────────────────────────────────────────────────
    ws.cell(r, 1, "INKOMSTEN").font = _font(bold=True, color=WHITE)
    ws.cell(r, 1).fill = _fill(C_GREEN)
    for i in range(n_months + 1):
        ws.cell(r, col_offset + i).fill = _fill(C_GREEN)
    ws.row_dimensions[r].height = 16
    r += 1

    inc_totals = {}
    for cat in inc_cats:
        if cat in ("ZZP Opname",):
            continue  # show separately as recommendation
        v_map = {}
        for m in ACTUAL_MONTHS:
            v_map[m] = pivot.loc[cat, m] if m in pivot.columns else 0.0
        avg = sum(v_map.values()) / 4
        for m in PROJ_MONTHS:
            v_map[m] = avg
        jaar = write_row(cat, v_map, PROJ_MONTHS, bg=C_ALT if list(inc_cats).index(cat) % 2 == 0 else None)
        inc_totals[cat] = jaar
        r += 1

    # ZZP Opname als aanbeveling
    zzp_map = {}
    for m in ACTUAL_MONTHS:
        zzp_map[m] = pivot.loc["ZZP Opname", m] if "ZZP Opname" in pivot.index and m in pivot.columns else 0.0
    for m in PROJ_MONTHS:
        zzp_map[m] = maand_opname_aanbevolen
    jaar = write_row("ZZP Opname  (aanbeveling Mei-Dec)", zzp_map, PROJ_MONTHS, bold=True)
    inc_totals["ZZP Opname"] = jaar
    r += 1

    # Inkomen subtotaal
    inc_sub = {}
    for m in ALL_MONTHS:
        inc_sub[m] = sum(inc_totals.get(c, 0) for c in inc_totals) / 12  # placeholder
    # Compute per month
    inc_monthly = {}
    for cat, jaar in inc_totals.items():
        vm = {}
        if cat == "ZZP Opname":
            vm = zzp_map
        else:
            for m in ACTUAL_MONTHS:
                vm[m] = pivot.loc[cat, m] if cat in pivot.index and m in pivot.columns else 0.0
            avg = sum(vm.values()) / 4
            for m in PROJ_MONTHS:
                vm[m] = avg
        for m in ALL_MONTHS:
            inc_monthly[m] = inc_monthly.get(m, 0.0) + vm.get(m, 0.0)

    inc_totaal_row = {}
    for m in ALL_MONTHS:
        inc_totaal_row[m] = inc_monthly[m]
    write_row("Totaal inkomsten", inc_totaal_row, PROJ_MONTHS, bold=True, bg=C_GREEN, color=WHITE)
    r += 1
    r += 1  # gap

    # ── KOSTEN ────────────────────────────────────────────────────────────────
    ws.cell(r, 1, "KOSTEN").font = _font(bold=True, color=WHITE)
    ws.cell(r, 1).fill = _fill("843C0C")
    for i in range(n_months + 1):
        ws.cell(r, col_offset + i).fill = _fill("843C0C")
    ws.row_dimensions[r].height = 16
    r += 1

    cost_monthly = {m: 0.0 for m in ALL_MONTHS}
    for idx, cat in enumerate(cost_cats):
        v_map = {}
        for m in ACTUAL_MONTHS:
            v_map[m] = pivot.loc[cat, m] if m in pivot.columns else 0.0
        avg = sum(v_map.values()) / 4
        for m in PROJ_MONTHS:
            v_map[m] = avg
        bg = C_ALT if idx % 2 == 0 else None
        write_row(cat, v_map, PROJ_MONTHS, bg=bg)
        for m in ALL_MONTHS:
            cost_monthly[m] = cost_monthly.get(m, 0.0) + v_map[m]
        r += 1

    write_row("Totaal kosten", cost_monthly, PROJ_MONTHS, bold=True, bg="FCE4D6")
    r += 1
    r += 1  # gap

    # ── NETTO ─────────────────────────────────────────────────────────────────
    netto_monthly = {m: inc_monthly[m] + cost_monthly[m] for m in ALL_MONTHS}
    write_row("NETTO  (inkomen + kosten)", netto_monthly, PROJ_MONTHS,
              bold=True, bg=C_SUB, color=C_DARK)
    r += 1

    # Legend
    r += 1
    ws.cell(r, 1, "Lichtblauw = werkelijk (Jan-Apr)").font = _font(color="595959", size=9)
    ws.cell(r, col_offset, " ").fill = _fill(C_ACTUAL)
    r += 1
    ws.cell(r, 1, "Geel = prognose (Mei-Dec, avg Jan-Apr)").font = _font(color="595959", size=9)
    ws.cell(r, col_offset, " ").fill = _fill(C_PROJ)


# ── Sheet 3: Samenvatting ─────────────────────────────────────────────────────

def write_samenvatting(ws, ods, knab_kosten_jaar, zzp_opname_reeds, abn_df):
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 36
    ws.freeze_panes = "A2"

    r = [1]

    def nxt():
        n = r[0]; r[0] += 1; return n

    def title(text):
        rn = nxt()
        ws.merge_cells(f"A{rn}:C{rn}")
        c = ws.cell(rn, 1, text)
        c.font = Font(bold=True, color=WHITE, name="Arial", size=13)
        c.fill = _fill(C_DARK)
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[rn].height = 26

    def section(text, bg=C_MED):
        rn = nxt()
        for col in (1, 2, 3):
            ws.cell(rn, col).fill = _fill(bg)
        ws.cell(rn, 1, text).font = _font(bold=True, color=WHITE)
        ws.row_dimensions[rn].height = 18

    def kv(label, amount, note="", alt=False, bold=False, color="000000", size=10):
        rn = nxt()
        bg = C_ALT if alt else None
        lc = ws.cell(rn, 1, "  " + label)
        lc.font = Font(bold=bold, color=color, name="Arial", size=size)
        if bg: lc.fill = _fill(bg)
        ac = ws.cell(rn, 2, amount)
        ac.number_format = EURO
        ac.alignment = Alignment(horizontal="right")
        ac.font = Font(bold=bold, color=color, name="Arial", size=size)
        if bg: ac.fill = _fill(bg)
        if note:
            nc = ws.cell(rn, 3, note)
            nc.font = _font(color="595959", size=9)
            if bg: nc.fill = _fill(bg)

    def sub(label, amount, bg=C_SUB, fg="000000", size=10):
        rn = nxt()
        for col in (1, 2):
            ws.cell(rn, col).fill = _fill(bg)
        lc = ws.cell(rn, 1, label)
        lc.font = Font(bold=True, color=fg, name="Arial", size=size)
        ac = ws.cell(rn, 2, amount)
        ac.number_format = EURO
        ac.alignment = Alignment(horizontal="right")
        ac.font = Font(bold=True, color=fg, name="Arial", size=size)
        if size > 10: ws.row_dimensions[rn].height = size + 8

    def gap():
        r[0] += 1

    # ── Totalen per jaar ───────────────────────────────────────────────────────
    ACTUAL_MONTHS = [f"2026-{m:02d}" for m in range(1, 5)]
    PROJ_MONTHS   = [f"2026-{m:02d}" for m in range(5, 13)]

    # Compute annual personal cost projection
    cost_act = abn_df[~abn_df["category"].isin(INCOME_CATS)]["total"].sum()
    avg_maand_kosten = cost_act / 4
    cost_prognose_jaar = cost_act + avg_maand_kosten * 8

    # Compute annual income (non-ZZP Opname) on ABN
    inc_act = abn_df[abn_df["category"].isin(INCOME_CATS) &
                     (abn_df["category"] != "ZZP Opname")]["total"].sum()
    avg_inc_maand = inc_act / 4
    inc_prognose_jaar = inc_act + avg_inc_maand * 8

    restant_opname = ods["beschikbaar"] - zzp_opname_reeds
    maand_opname  = restant_opname / 8
    zzp_jaar_abn  = zzp_opname_reeds + restant_opname

    totaal_ink_jaar  = inc_prognose_jaar + zzp_jaar_abn
    netto_jaar       = totaal_ink_jaar + cost_prognose_jaar
    maand_kosten     = avg_maand_kosten

    title(f"Samenvatting Prognose {YEAR}")
    gap()

    section("ODS — ZAKELIJK")
    kv("Omzet excl. BTW", ods["omzet"], note="Opgegeven", alt=True, bold=True)
    kv("Bedrijfskosten", -ods["kosten_ods"], note="Werkelijk Jan-Apr + prognose Mei-Dec")
    kv("Winst uit onderneming", ods["winst"])
    kv("IB te betalen  (schatting)", -ods["ib_te_betalen"],
       note="Eenmanszaak, Box 1  — RESERVEREN", alt=True)
    sub("Beschikbaar voor privé  (netto)", ods["beschikbaar"], bg=C_GREEN, fg=WHITE, size=11)
    gap()

    kv("BTW reservering  (kwartaal)", -(ods["btw_reservering"] / 4),
       note="Per kwartaal apart zetten in ODS/Knab")
    kv("IB reservering  (maand)", -(ods["ib_te_betalen"] / 12),
       note="Maandelijks apart zetten")
    gap()

    section("ZZP OPNAME — ODS → ABN")
    kv("Reeds opgenomen  (Jan-Apr)", -zzp_opname_reeds, note="Werkelijk", alt=True)
    kv("Resterende opnameruimte  (Mei-Dec)", restant_opname)
    sub("→ Aanbevolen opname per maand", maand_opname, bg=C_SUB)
    gap()

    section("ABN PRIVÉ — INKOMSTEN EN KOSTEN")
    kv("ZZP Opname  (prognose heel jaar)", zzp_jaar_abn, note="Incl. reeds opgenomen", alt=True)
    kv("Salaris + overig inkomen  (prognose)", inc_prognose_jaar,
       note=f"Avg. {avg_inc_maand:.0f}/mnd × 12")
    sub("Totaal inkomsten ABN  (prognose)", totaal_ink_jaar)
    gap()
    kv("Kosten  (werkelijk Jan-Apr)", cost_act, note=f"Avg. {avg_maand_kosten:.0f}/mnd")
    kv("Kosten  (prognose Mei-Dec)", avg_maand_kosten * 8, alt=True,
       note=f"Avg. {avg_maand_kosten:.0f}/mnd × 8")
    sub("Totaal kosten ABN  (prognose)", cost_prognose_jaar, bg="FCE4D6")
    gap()

    netto_color = C_GREEN if netto_jaar >= 0 else C_RED
    sub("NETTO VRIJ BESTEEDBAAR  (prognose 2026)", netto_jaar,
        bg=netto_color, fg=WHITE, size=12)
    kv("→ Gemiddeld per maand", netto_jaar / 12,
       note="Inkomsten - kosten (excl. IB reservering)")
    gap()

    section("MAANDELIJKSE AANBEVELING  (Mei–Dec 2026)")
    kv("ZZP Opname van ODS naar ABN", maand_opname,
       note="Maximaal opneembaar na belasting", alt=True, bold=True)
    kv("Maandelijkse kosten  (prognose)", maand_kosten,
       note="Gemiddelde Jan-Apr × 12")
    kv("Salaris + overig  (prognose)", avg_inc_maand, alt=True,
       note="Automatisch / vaste inkomsten")
    sub("Spaarruimte per maand", maand_opname + avg_inc_maand + maand_kosten,
        bg=C_GREEN, fg=WHITE, size=11)
    gap()
    kv("IB reservering  (maandelijks apart)", -(ods["ib_te_betalen"] / 12),
       note="Spaarrekening — voor aanslag 2026", alt=True)
    kv("BTW reservering  (maandelijks apart)", -(ods["btw_reservering"] / 12),
       note="Naar Knab/spaar — per kwartaal betalen")


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    print(f"\nPrognose {YEAR}")
    print("=" * 50)

    # Load DB data
    abn_df = load_abn_2026()
    knab_df = load_knab_2026()

    ACTUAL_MONTHS = [f"2026-{m:02d}" for m in range(1, 5)]

    # Knab kosten Jan-Apr
    knab_werkelijk = abs(knab_df["total"].sum()) if not knab_df.empty else 0.0
    knab_prognose  = knab_werkelijk / 4 * 8   # project Mei-Dec
    knab_kosten_jaar = knab_werkelijk + knab_prognose

    # ZZP opname reeds genomen (ABN)
    zzp_mask = (abn_df["category"] == "ZZP Opname")
    zzp_opname_reeds = abn_df[zzp_mask]["total"].sum() if zzp_mask.any() else 0.0

    # Salaris / zorgtoeslag monthly avg (for display)
    sal_act  = abn_df[abn_df["category"] == "Salaris"]["total"].sum()
    zorg_act = abn_df[abn_df["category"] == "Zorgtoeslag"]["total"].sum()
    avg_sal  = sal_act / 4
    avg_zorg = zorg_act / 4

    # ODS fiscal calculation
    ods = bereken_ods(OMZET_ODS, knab_kosten_jaar)

    print(f"  Omzet ODS:              {OMZET_ODS:>10,.0f}")
    print(f"  Knab kosten (jaar):    -{knab_kosten_jaar:>10,.0f}  (werkelijk {knab_werkelijk:.0f} + prognose {knab_prognose:.0f})")
    print(f"  Winst:                  {ods['winst']:>10,.0f}")
    print(f"  Belastbare winst:       {ods['belastbare_winst']:>10,.0f}")
    print(f"  IB te betalen:         -{ods['ib_te_betalen']:>10,.0f}")
    print(f"  Beschikbaar privé:      {ods['beschikbaar']:>10,.0f}")
    print(f"  ZZP opname reeds:      -{zzp_opname_reeds:>10,.0f}")
    print(f"  Restant op te nemen:    {ods['beschikbaar']-zzp_opname_reeds:>10,.0f}  ({(ods['beschikbaar']-zzp_opname_reeds)/8:.0f}/mnd)")
    print()

    cost_act = abn_df[~abn_df["category"].isin(INCOME_CATS)]["total"].sum()
    print(f"  ABN kosten Jan-Apr:     {cost_act:>10,.0f}  (avg {cost_act/4:.0f}/mnd)")
    print(f"  ABN kosten prognose:    {cost_act + cost_act/4*8:>10,.0f}  (heel 2026)")

    maand_opname = (ods["beschikbaar"] - zzp_opname_reeds) / 8

    # Write Excel
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "ODS Berekening"
    write_ods_sheet(ws1, ods, knab_werkelijk, knab_prognose, zzp_opname_reeds)

    ws2 = wb.create_sheet("ABN Prognose")
    write_abn_sheet(ws2, abn_df, maand_opname, avg_sal, avg_zorg)

    ws3 = wb.create_sheet("Samenvatting")
    write_samenvatting(ws3, ods, knab_kosten_jaar, zzp_opname_reeds, abn_df)

    out = Path(OUTPUT_DIR) / f"prognose_{YEAR}.xlsx"
    wb.save(out)
    print(f"\n  Opgeslagen: {out}\n")


if __name__ == "__main__":
    main()
