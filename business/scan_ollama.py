import sys
sys.stdout.reconfigure(encoding="utf-8")

import json
import base64
import time
from pathlib import Path

import requests
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from config import (
    INVOICE_DIRS,
    SUPPORTED_EXTENSIONS,
    OUTPUT_DIR,
    OLLAMA_URL,
    OLLAMA_MODEL,
    OLLAMA_TIMEOUT,
)

PROMPT = """Je bent een Nederlandse boekhoud-assistent.
Analyseer dit bonnetje of factuur zorgvuldig.
Geef ALLEEN een geldig JSON object terug, geen uitleg, geen markdown, geen tekst ervoor of erna.

JSON formaat (gebruik exact deze sleutels):
{
  "vendor": "naam van de leverancier of winkel",
  "invoice_number": "factuurnummer of bonnetnummer, of null als niet zichtbaar",
  "date": "datum in formaat YYYY-MM-DD, of null als niet leesbaar",
  "amount_excl_btw": 0.00,
  "btw_percentage": 21,
  "btw_amount": 0.00,
  "amount_incl_btw": 0.00,
  "currency": "EUR",
  "category": "kies uit: Software, Hardware, Kantoor, Reiskosten, Marketing, Telecom, Professionele Diensten, Maaltijden, Overig",
  "description": "korte omschrijving in het Nederlands van wat er gekocht is",
  "payment_method": "kies uit: Bank, iDEAL, Cash, CreditCard, Onbekend",
  "confidence": 0.85
}

Confidence = hoe zeker je bent van de extractie (0.0 = totaal onzeker, 1.0 = volledig zeker).
Als een bedrag niet leesbaar is, zet dan 0.00. Verzin nooit bedragen.
Als datum niet leesbaar is, zet null."""

COLUMNS = [
    ("Bestand",          18, None),
    ("Leverancier",      22, None),
    ("Factuurnr",        16, None),
    ("Datum",            12, None),
    ("Excl. BTW",        12, '€ #,##0.00;-€ #,##0.00'),
    ("BTW%",              7, '0"%"'),
    ("BTW",              11, '€ #,##0.00;-€ #,##0.00'),
    ("Incl. BTW",        12, '€ #,##0.00;-€ #,##0.00'),
    ("Valuta",            8, None),
    ("Categorie",        20, None),
    ("Omschrijving",     35, None),
    ("Betaalwijze",      14, None),
    ("Confidence",       12, '0.00'),
    ("Verwerkingstijd",  16, '0.0" s"'),
    ("Model",            16, None),
]

HDR_FILL  = PatternFill("solid", fgColor="1F4E79")
HDR_FONT  = Font(bold=True, color="FFFFFF", name="Arial", size=10)
ROW_FONT  = Font(name="Arial", size=10)
ALT_FILL  = PatternFill("solid", fgColor="F2F2F2")
THIN      = Side(style="thin", color="D9D9D9")
ROW_BORDER = Border(bottom=Side(style="thin", color="E0E0E0"))


def check_ollama():
    try:
        r = requests.get("http://localhost:11434/api/tags", timeout=5)
        r.raise_for_status()
        models = [m["name"].split(":")[0] for m in r.json().get("models", [])]
        if OLLAMA_MODEL not in models:
            print(f"❌ llava model niet gevonden. Download het eerst:\n   ollama pull llava")
            sys.exit(1)
    except requests.exceptions.ConnectionError:
        print(
            "❌ Ollama reageert niet op http://localhost:11434\n"
            "   Start Ollama eerst:\n"
            "   - Open de Ollama app in je systray (rechtsonderin de taakbalk)\n"
            "   - Of run in een terminal: ollama serve"
        )
        sys.exit(1)


def collect_files(year_dirs: dict) -> list[tuple[str, Path]]:
    files = []
    for year, dir_str in year_dirs.items():
        d = Path(dir_str)
        if not d.exists():
            print(f"⚠️  Map niet gevonden: {d}")
            continue
        images = [p for p in d.rglob("*") if p.suffix.lower() in SUPPORTED_EXTENSIONS]
        pdfs   = [p for p in d.rglob("*") if p.suffix.lower() == ".pdf"]
        for p in pdfs:
            print(f"⚠️  {p.name} overgeslagen — llava ondersteunt geen PDFs")
        print(f"📂 {year} — {len(images)} afbeelding(en) gevonden")
        files.extend((year, p) for p in sorted(images))
    return files


def process_image(filepath: Path) -> dict:
    with open(filepath, "rb") as f:
        image_b64 = base64.b64encode(f.read()).decode("utf-8")

    payload = {
        "model": OLLAMA_MODEL,
        "prompt": PROMPT,
        "images": [image_b64],
        "stream": False,
        "format": "json",
    }

    start = time.time()
    response = requests.post(OLLAMA_URL, json=payload, timeout=OLLAMA_TIMEOUT)
    elapsed = round(time.time() - start, 1)

    raw = response.json()["response"].strip()
    if raw.startswith("```"):
        raw = raw.split("```")[1]
        if raw.startswith("json"):
            raw = raw[4:]
    data = json.loads(raw.strip())
    data["filename"]          = filepath.name
    data["model"]             = f"ollama/{OLLAMA_MODEL}"
    data["processing_time_s"] = elapsed
    return data


def print_result(data: dict):
    vendor   = data.get("vendor") or "?"
    cat      = data.get("category") or "?"
    excl     = data.get("amount_excl_btw") or 0.0
    btw      = data.get("btw_amount") or 0.0
    incl     = data.get("amount_incl_btw") or 0.0
    conf     = data.get("confidence") or 0.0
    elapsed  = data.get("processing_time_s", 0)
    print(
        f"  ✅ {data['filename']} — {vendor} | {cat} | "
        f"€{excl:.2f} + €{btw:.2f} BTW = €{incl:.2f} | "
        f"conf: {conf:.2f} ({elapsed}s)"
    )


def write_excel(results: list[dict], year_label: str):
    def sort_key(r):
        d = r.get("date")
        return (0, d) if d else (1, "")

    results.sort(key=sort_key)

    wb = Workbook()
    ws = wb.active
    ws.title = "Facturen"
    ws.freeze_panes = "A2"

    for c, (header, width, _) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=c, value=header)
        cell.fill   = HDR_FILL
        cell.font   = HDR_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(c)].width = width

    ws.row_dimensions[1].height = 20

    field_map = [
        "filename", "vendor", "invoice_number", "date",
        "amount_excl_btw", "btw_percentage", "btw_amount", "amount_incl_btw",
        "currency", "category", "description", "payment_method",
        "confidence", "processing_time_s", "model",
    ]

    for r, row_data in enumerate(results, 2):
        fill = ALT_FILL if r % 2 == 0 else None
        for c, (field, (_, _, fmt)) in enumerate(zip(field_map, COLUMNS), 1):
            val  = row_data.get(field)
            cell = ws.cell(row=r, column=c, value=val)
            cell.font      = ROW_FONT
            cell.border    = ROW_BORDER
            cell.alignment = Alignment(vertical="center")
            if fill:
                cell.fill = fill
            if fmt:
                cell.number_format = fmt

    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"

    out_path = OUTPUT_DIR / f"facturen_{year_label}.xlsx"
    wb.save(out_path)
    return out_path


def main():
    check_ollama()

    arg = sys.argv[1] if len(sys.argv) > 1 else "2026"

    if arg == "all":
        year_dirs   = INVOICE_DIRS
        year_label  = "all"
    elif arg in INVOICE_DIRS:
        year_dirs  = {arg: INVOICE_DIRS[arg]}
        year_label = arg
    else:
        print(f"❌ Jaar '{arg}' niet gevonden. Beschikbaar: {', '.join(INVOICE_DIRS)}")
        sys.exit(1)

    files = collect_files(year_dirs)
    if not files:
        print("Geen afbeeldingen gevonden.")
        sys.exit(0)

    print(f"🔄 Verbinden met Ollama {OLLAMA_MODEL} model...")

    results = []
    errors  = 0
    for year, filepath in files:
        try:
            data = process_image(filepath)
            results.append(data)
            print_result(data)
        except Exception as e:
            errors += 1
            results.append({"filename": filepath.name, "error": str(e)})
            print(f"  ❌ {filepath.name} — {e}")

    out_path = write_excel(results, year_label)

    ok = len(results) - errors
    print(f"\n✅ Klaar — {ok} verwerkt, {errors} fout(en)")
    print(f"📁 Opgeslagen: {out_path}")


if __name__ == "__main__":
    main()
