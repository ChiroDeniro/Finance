"""
Saldo Rekeningen — actueel saldo per rekening
==============================================
Leest automatisch uit alle exports:
  ABN AMRO : TAB-bestanden (balance_after van laatste transactie)
  Revolut  : CSV (Saldo kolom, laatste VOLTOOID rij)
  Knab     : handmatig via ODS_SALDO in networth_config.py
  DeGiro   : netto kasstroom uit Transactions.csv (geen realtime koers)
  Bitvavo  : EUR saldo berekend uit deposit/buy/sell transacties

Run:
    cd business
    python saldo_rekeningen.py
"""

import sys
sys.stdout.reconfigure(encoding="utf-8")

from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

from networth_config import ABN_LABELS, ODS_SALDO, PEILDATUM
from config import OUTPUT_DIR

INPUT_DIR = Path(__file__).parent.parent / "bookkeeping" / "input"

EURO    = "€ #,##0.00;-€ #,##0.00"
C_DARK  = "1F4E79"
C_MED   = "2E75B6"
C_ALT   = "F2F2F2"
C_GREEN = "375623"
WHITE   = "FFFFFF"


def _fill(hex_c):
    return PatternFill("solid", fgColor=hex_c)


def _font(bold=False, color="000000", size=10):
    return Font(bold=bold, color=color, name="Arial", size=size)


def _read_abn_balances():
    """Laatste balance_after per rekeningnummer uit alle TAB-bestanden."""
    balances = {}
    for f in INPUT_DIR.glob("**/*.TAB"):
        try:
            with open(f, encoding="latin-1") as fh:
                for line in fh:
                    parts = line.strip().split("\t")
                    if len(parts) < 5:
                        continue
                    account, _, date_str, _, bal_str = parts[:5]
                    account  = account.strip()
                    date_str = date_str.strip()
                    bal      = float(bal_str.strip().replace(",", "."))
                    if account not in balances or date_str > balances[account][0]:
                        balances[account] = (date_str, bal)
        except Exception:
            continue
    return {acc: (bal, date_str) for acc, (date_str, bal) in balances.items()}


def _read_degiro_netto():
    """Netto kasstroom uit DeGiro Transactions.csv (som Totaal EUR)."""
    degiro_dir = INPUT_DIR / "DeGiro"
    if not degiro_dir.exists():
        return None, None
    total = 0.0
    last_date = None
    found = False
    for f in degiro_dir.glob("*.csv"):
        try:
            raw = pd.read_csv(f, encoding="utf-8-sig", dtype=str, header=0)
            raw.columns = [str(c).strip() for c in raw.columns]
            if "Totaal EUR" not in raw.columns or len(raw.columns) < 16:
                continue
            if "Datum" not in raw.columns:
                continue
            # Parse dates (DD-MM-YYYY)
            dates = pd.to_datetime(raw.iloc[:, 0].str.strip(), format="%d-%m-%Y", errors="coerce")
            totaal = (raw["Totaal EUR"].astype(str).str.strip()
                      .str.replace(",", ".", regex=False)
                      .apply(lambda x: float(x) if x not in ("", "nan") else 0.0))
            total += totaal.sum()
            d = dates.dropna().max()
            if d is not pd.NaT and (last_date is None or d > last_date):
                last_date = d
            found = True
        except Exception:
            continue
    if not found:
        return None, None
    return total, last_date


def _read_bitvavo_eur_balance():
    """EUR saldo in Bitvavo = som van alle EUR-flows (deposits + sells - buys)."""
    bitvavo_dir = INPUT_DIR / "Bitvavo - crypto"
    if not bitvavo_dir.exists():
        return None, None
    total = 0.0
    last_date = None
    found = False
    for f in bitvavo_dir.glob("*.csv"):
        try:
            raw = pd.read_csv(f, encoding="utf-8-sig", dtype=str)
            raw.columns = raw.columns.str.strip()
            if not {"Date", "Time", "Type", "Currency", "Amount", "Status"}.issubset(raw.columns):
                continue
            raw = raw[raw["Status"].str.lower() == "completed"].copy()
            if raw.empty:
                continue
            dates = pd.to_datetime(
                raw["Date"] + " " + raw["Time"].str[:8], format="mixed", errors="coerce"
            )
            def _num(s):
                return (s.fillna("0").astype(str).str.replace(",", ".", regex=False)
                        .apply(lambda x: float(x) if x not in ("", "nan") else 0.0))
            rpa = _num(raw.get("Received / Paid Amount", pd.Series("0", index=raw.index)))
            amt = _num(raw["Amount"])
            typ = raw["Type"].str.lower()
            cur = raw["Currency"].fillna("")
            # EUR deposits: rpa is 0, use Amount
            is_eur_dep = (typ == "deposit") & (cur == "EUR") & (rpa == 0)
            eur_flow = rpa.copy()
            eur_flow[is_eur_dep] = amt[is_eur_dep]
            total += eur_flow.sum()
            d = dates.dropna().max()
            if d is not pd.NaT and (last_date is None or d > last_date):
                last_date = d
            found = True
        except Exception:
            continue
    if not found:
        return None, None
    return total, last_date


def _read_revolut_balance():
    """Laatste saldo uit Revolut CSV-bestanden (Saldo kolom)."""
    revolut_dir = INPUT_DIR / "Revolut"
    if not revolut_dir.exists():
        return None, None

    latest_date = None
    latest_bal  = None

    for f in revolut_dir.glob("*.csv"):
        try:
            raw = pd.read_csv(f, encoding="utf-8-sig", dtype=str)
            raw.columns = raw.columns.str.strip()
            if not {"Saldo", "Status", "Datum voltooid"}.issubset(raw.columns):
                continue
            voltooid = raw[raw["Status"].str.upper() == "VOLTOOID"].copy()
            voltooid["_date"] = pd.to_datetime(
                voltooid["Datum voltooid"], format="%Y-%m-%d %H:%M:%S", errors="coerce"
            )
            voltooid = voltooid.dropna(subset=["_date"]).sort_values("_date")
            if voltooid.empty:
                continue
            last    = voltooid.iloc[-1]
            bal     = float(str(last["Saldo"]).replace(",", "."))
            date_v  = last["_date"]
            if latest_date is None or date_v > latest_date:
                latest_date = date_v
                latest_bal  = bal
        except Exception:
            continue

    return latest_bal, latest_date


def _write_sheet(ws, accounts):
    """
    accounts: list of (naam, bedrag_or_None, valuta, datum_str)
    Schrijft tabel naar ws, retourneert totaal EUR.
    """
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 18
    ws.freeze_panes = "A2"

    for c, h in enumerate(["Rekening", "Saldo", "Valuta", "Stand per"], 1):
        cell = ws.cell(1, c, h)
        cell.fill = _fill(C_DARK)
        cell.font = _font(bold=True, color=WHITE)
        cell.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 20

    totaal = 0.0
    for i, (naam, bedrag, valuta, datum) in enumerate(accounts, 2):
        bg = _fill(C_ALT) if i % 2 == 0 else None

        def _cell(col, val, fmt=None, align=None):
            c = ws.cell(i, col, val)
            c.font = _font()
            if bg:
                c.fill = bg
            if fmt:
                c.number_format = fmt
            if align:
                c.alignment = Alignment(horizontal=align)
            return c

        _cell(1, naam)
        _cell(2, bedrag if bedrag is not None else 0.0, fmt=EURO, align="right")
        _cell(3, valuta, align="center")
        _cell(4, datum or "—", align="center")

        if bedrag is not None and valuta == "EUR":
            totaal += bedrag

    r = len(accounts) + 2
    for col in range(1, 5):
        ws.cell(r, col).fill = _fill(C_GREEN)
    ws.cell(r, 1, "TOTAAL SALDO (EUR)").font = _font(bold=True, color=WHITE)
    ws.cell(r, 1).fill = _fill(C_GREEN)
    tc = ws.cell(r, 2, totaal)
    tc.font      = _font(bold=True, color=WHITE)
    tc.fill      = _fill(C_GREEN)
    tc.number_format = EURO
    tc.alignment = Alignment(horizontal="right")
    ws.row_dimensions[r].height = 20

    return totaal


def main():
    print("\nSaldo Rekeningen")
    print("=" * 40)

    abn                       = _read_abn_balances()
    revolut_bal, revolut_date = _read_revolut_balance()
    degiro_netto, degiro_date = _read_degiro_netto()
    bitvavo_eur, bitvavo_date = _read_bitvavo_eur_balance()

    accounts = []

    for acc_nr in ["536542171", "844835730"]:
        naam = ABN_LABELS.get(acc_nr, f"ABN AMRO {acc_nr}")
        if acc_nr in abn:
            bal, date_str = abn[acc_nr]
            datum = f"{date_str[:4]}-{date_str[4:6]}-{date_str[6:]}" if len(date_str) == 8 else date_str
            accounts.append((naam, bal, "EUR", datum))
            print(f"  {naam:<38} €{bal:>10,.2f}  ({datum})")
        else:
            accounts.append((naam, None, "EUR", "niet gevonden"))
            print(f"  {naam:<38}  — geen TAB bestand gevonden")

    knab_datum = f"{PEILDATUM} (handmatig)"
    accounts.append(("Knab Zakelijke Rekening", ODS_SALDO, "EUR", knab_datum))
    if ODS_SALDO:
        print(f"  {'Knab Zakelijke Rekening':<38} €{ODS_SALDO:>10,.2f}  ({knab_datum})")
    else:
        print(f"  {'Knab Zakelijke Rekening':<38}  €0,00 — vul ODS_SALDO in networth_config.py")

    if revolut_bal is not None:
        rev_datum = revolut_date.strftime("%Y-%m-%d") if revolut_date else PEILDATUM
        accounts.append(("Revolut", revolut_bal, "EUR", rev_datum))
        print(f"  {'Revolut':<38} €{revolut_bal:>10,.2f}  ({rev_datum})")
    else:
        accounts.append(("Revolut", None, "EUR", "geen CSV gevonden"))
        print(f"  {'Revolut':<38}  — geen CSV gevonden in input/Revolut/")

    # DeGiro: toon netto kasstroom uit trades (geen realtime portefeuille waarde)
    if degiro_netto is not None:
        dg_datum = degiro_date.strftime("%Y-%m-%d") if degiro_date else PEILDATUM
        accounts.append(("DeGiro (netto kasstroom trades)", degiro_netto, "EUR", dg_datum))
        print(f"  {'DeGiro (netto kasstroom trades)':<38} €{degiro_netto:>10,.2f}  ({dg_datum})")
        print(f"  {'  ↳ realtime waarde: zie BELEGGINGEN in networth_config.py':<38}")
    else:
        accounts.append(("DeGiro", None, "EUR", "geen CSV gevonden"))
        print(f"  {'DeGiro':<38}  — geen CSV gevonden in input/DeGiro/")

    # Bitvavo: EUR saldo (gestort - uitgegeven aan crypto + ontvangen van verkopen)
    if bitvavo_eur is not None:
        bv_datum = bitvavo_date.strftime("%Y-%m-%d") if bitvavo_date else PEILDATUM
        accounts.append(("Bitvavo (EUR kassaldo)", bitvavo_eur, "EUR", bv_datum))
        print(f"  {'Bitvavo (EUR kassaldo)':<38} €{bitvavo_eur:>10,.2f}  ({bv_datum})")
    else:
        accounts.append(("Bitvavo", None, "EUR", "geen CSV gevonden"))
        print(f"  {'Bitvavo':<38}  — geen CSV gevonden in input/Bitvavo - crypto/")

    totaal = sum(b for _, b, v, _ in accounts if b is not None and v == "EUR")
    print(f"\n  {'TOTAAL SALDO (EUR)':<38} €{totaal:>10,.2f}")

    wb = Workbook()
    ws = wb.active
    ws.title = "Saldo Rekeningen"
    _write_sheet(ws, accounts)

    out = Path(OUTPUT_DIR) / f"saldo_rekeningen_{PEILDATUM}.xlsx"
    wb.save(out)
    print(f"\n  Opgeslagen: {out}\n")


if __name__ == "__main__":
    main()
