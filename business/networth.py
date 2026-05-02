import sys
sys.stdout.reconfigure(encoding="utf-8")

from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from networth_config import (
    PEILDATUM, ABN_LABELS, ODS_SALDO,
    BELEGGINGEN, ACTIVA, SCHULDEN, VORDERINGEN_OP_ODS,
)
from config import OUTPUT_DIR

EURO = '€ #,##0.00;-€ #,##0.00'
C_DARK  = "1F4E79"
C_MED   = "2E75B6"
C_SUB   = "BDD7EE"
C_GREEN = "375623"
C_ALT   = "F2F2F2"

INPUT_DIR = Path(__file__).parent.parent / "bookkeeping" / "input"


def _fill(hex_c):
    return PatternFill("solid", fgColor=hex_c)


def _font(bold=False, color="000000", size=10):
    return Font(bold=bold, color=color, name="Arial", size=size)


def read_abn_balances():
    balances = {}
    for f in INPUT_DIR.glob("**/*.TAB"):
        try:
            with open(f, encoding="latin-1") as fh:
                for line in fh:
                    parts = line.strip().split("\t")
                    if len(parts) < 5:
                        continue
                    account, _, date_str, _, bal_str = parts[:5]
                    account = account.strip()
                    date_str = date_str.strip()
                    bal = float(bal_str.strip().replace(",", "."))
                    if account not in balances or date_str > balances[account][0]:
                        balances[account] = (date_str, bal)
        except Exception:
            continue
    return {acc: bal for acc, (_, bal) in balances.items()}


def _read_revolut_balance():
    revolut_dir = INPUT_DIR / "Revolut"
    if not revolut_dir.exists():
        return None, None
    latest_date = None
    latest_bal  = None
    for f in revolut_dir.glob("*.csv"):
        try:
            raw = pd.read_csv(f, encoding="utf-8-sig", dtype=str)
            raw.columns = raw.columns.str.strip()
            if not {"Saldo", "Status", "Datum voltooid"}.issubset(raw.columns):
                continue
            voltooid = raw[raw["Status"].str.upper() == "VOLTOOID"].copy()
            voltooid["_date"] = pd.to_datetime(
                voltooid["Datum voltooid"], format="%Y-%m-%d %H:%M:%S", errors="coerce"
            )
            voltooid = voltooid.dropna(subset=["_date"]).sort_values("_date")
            if voltooid.empty:
                continue
            last   = voltooid.iloc[-1]
            bal    = float(str(last["Saldo"]).replace(",", "."))
            date_v = last["_date"]
            if latest_date is None or date_v > latest_date:
                latest_date = date_v
                latest_bal  = bal
        except Exception:
            continue
    return latest_bal, latest_date


def _read_degiro_netto():
    degiro_dir = INPUT_DIR / "DeGiro"
    if not degiro_dir.exists():
        return None, None
    total = 0.0
    last_date = None
    found = False
    for f in degiro_dir.glob("*.csv"):
        try:
            raw = pd.read_csv(f, encoding="utf-8-sig", dtype=str, header=0)
            raw.columns = [str(c).strip() for c in raw.columns]
            if "Totaal EUR" not in raw.columns or "Datum" not in raw.columns:
                continue
            dates = pd.to_datetime(raw.iloc[:, 0].str.strip(), format="%d-%m-%Y", errors="coerce")
            totaal = (raw["Totaal EUR"].astype(str).str.strip()
                      .str.replace(",", ".", regex=False)
                      .apply(lambda x: float(x) if x not in ("", "nan") else 0.0))
            total += totaal.sum()
            d = dates.dropna().max()
            if d is not pd.NaT and (last_date is None or d > last_date):
                last_date = d
            found = True
        except Exception:
            continue
    return (total, last_date) if found else (None, None)


def _read_bitvavo_eur_balance():
    bitvavo_dir = INPUT_DIR / "Bitvavo - crypto"
    if not bitvavo_dir.exists():
        return None, None
    total = 0.0
    last_date = None
    found = False
    for f in bitvavo_dir.glob("*.csv"):
        try:
            raw = pd.read_csv(f, encoding="utf-8-sig", dtype=str)
            raw.columns = raw.columns.str.strip()
            if not {"Date", "Time", "Type", "Currency", "Amount", "Status"}.issubset(raw.columns):
                continue
            raw = raw[raw["Status"].str.lower() == "completed"].copy()
            if raw.empty:
                continue
            dates = pd.to_datetime(
                raw["Date"] + " " + raw["Time"].str[:8], format="mixed", errors="coerce"
            )
            def _num(s):
                return (s.fillna("0").astype(str).str.replace(",", ".", regex=False)
                        .apply(lambda x: float(x) if x not in ("", "nan") else 0.0))
            rpa      = _num(raw.get("Received / Paid Amount", pd.Series("0", index=raw.index)))
            amt      = _num(raw["Amount"])
            typ      = raw["Type"].str.lower()
            cur      = raw["Currency"].fillna("")
            is_eur_dep = (typ == "deposit") & (cur == "EUR") & (rpa == 0)
            eur_flow = rpa.copy()
            eur_flow[is_eur_dep] = amt[is_eur_dep]
            total += eur_flow.sum()
            d = dates.dropna().max()
            if d is not pd.NaT and (last_date is None or d > last_date):
                last_date = d
            found = True
        except Exception:
            continue
    return (total, last_date) if found else (None, None)


def write_vermogen(ws, abn_balances, revolut_bal=None, revolut_date=None,
                   degiro_netto=None, degiro_date=None,
                   bitvavo_eur=None, bitvavo_date=None):
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 16
    ws.freeze_panes = "A2"
    r = [1]

    def row():
        n = r[0]; r[0] += 1; return n

    def title(text):
        rn = row()
        c = ws.cell(rn, 1, text)
        c.font = Font(bold=True, color="FFFFFF", name="Arial", size=12)
        c.fill = _fill(C_DARK)
        c.alignment = Alignment(vertical="center")
        ws.merge_cells(f"A{rn}:B{rn}")
        ws.row_dimensions[rn].height = 24

    def section(text):
        rn = row()
        for col in (1, 2):
            ws.cell(rn, col).fill = _fill(C_MED)
        c = ws.cell(rn, 1, text)
        c.font = _font(bold=True, color="FFFFFF")

    def item(label, amount, alt=False):
        rn = row()
        lc = ws.cell(rn, 1, f"  {label}")
        ac = ws.cell(rn, 2, amount if amount is not None else 0.0)
        for cell in (lc, ac):
            cell.font = _font()
            if alt:
                cell.fill = _fill(C_ALT)
        ac.number_format = EURO
        ac.alignment = Alignment(horizontal="right")
        return amount if amount is not None else 0.0

    def subtotal(label, total, color=C_SUB, text_color="000000", size=10):
        rn = row()
        lc = ws.cell(rn, 1, label)
        ac = ws.cell(rn, 2, total)
        for cell in (lc, ac):
            cell.font = Font(bold=True, color=text_color, name="Arial", size=size)
            cell.fill = _fill(color)
        ac.number_format = EURO
        ac.alignment = Alignment(horizontal="right")
        if size > 10:
            ws.row_dimensions[rn].height = size + 10
        return total

    def gap():
        row()

    title(f"Netto Vermogen — {PEILDATUM}")
    gap()

    total_bezittingen = 0

    # ── Privé Bankrekeningen ───────────────────────────────────────────────────
    section("Privé Bankrekeningen  (automatisch uit TAB / CSV export)")
    bank_total = 0.0
    i = 0
    for acc_nr in ["536542171", "844835730"]:
        label = ABN_LABELS.get(acc_nr, f"ABN AMRO {acc_nr}")
        bal   = abn_balances.get(acc_nr)
        if bal is None:
            bank_total += item(f"{label}  (niet gevonden)", 0.0, i % 2 == 0)
        else:
            bank_total += item(label, bal, i % 2 == 0)
        i += 1

    if revolut_bal is not None:
        rev_label = "Revolut"
        if revolut_date:
            rev_label += f"  (stand {revolut_date.strftime('%Y-%m-%d')})"
        bank_total += item(rev_label, revolut_bal, i % 2 == 0)
        i += 1
    else:
        item("Revolut  (geen CSV gevonden)", 0.0, i % 2 == 0)
        i += 1

    subtotal("Subtotaal privé bank", bank_total)
    total_bezittingen += bank_total
    gap()

    # ── Zakelijke Rekening ────────────────────────────────────────────────────
    section("Zakelijke Rekening (Knab)")
    knab_label = "Knab Zakelijke Rekening"
    if ODS_SALDO == 0:
        knab_label += "  ← vul ODS_SALDO in networth_config.py"
    ods = item(knab_label, ODS_SALDO)
    subtotal("Subtotaal zakelijk", ods)
    total_bezittingen += ods
    gap()

    # ── Beleggingen ───────────────────────────────────────────────────────────
    section("Beleggingen")
    beleg = 0.0
    i = 0
    for k, v in BELEGGINGEN.items():
        beleg += item(k, v, i % 2 == 0)
        i += 1

    if degiro_netto is not None:
        dg_label = "DeGiro  (netto kasstroom trades)"
        if degiro_date:
            dg_label += f"  — t/m {degiro_date.strftime('%Y-%m-%d')}"
        beleg += item(dg_label, degiro_netto, i % 2 == 0)
        i += 1
    else:
        item("DeGiro  (geen CSV gevonden in input/DeGiro/)", 0.0, i % 2 == 0)
        i += 1

    if bitvavo_eur is not None:
        bv_label = "Bitvavo  (EUR kassaldo)"
        if bitvavo_date:
            bv_label += f"  — t/m {bitvavo_date.strftime('%Y-%m-%d')}"
        beleg += item(bv_label, bitvavo_eur, i % 2 == 0)
        i += 1
    else:
        item("Bitvavo  (geen CSV gevonden in input/Bitvavo - crypto/)", 0.0, i % 2 == 0)
        i += 1

    subtotal("Subtotaal beleggingen", beleg)
    total_bezittingen += beleg
    gap()

    # ── Activa ────────────────────────────────────────────────────────────────
    section("Activa (geschatte marktwaarde)")
    activa = sum(item(k, v, i % 2 == 0) for i, (k, v) in enumerate(ACTIVA.items()))
    subtotal("Subtotaal activa", activa)
    total_bezittingen += activa
    gap()

    subtotal("TOTAAL BEZITTINGEN", total_bezittingen, color=C_GREEN, text_color="FFFFFF", size=11)
    gap()

    # ── Schulden ──────────────────────────────────────────────────────────────
    section("Schulden")
    schulden = sum(item(k, v, i % 2 == 0) for i, (k, v) in enumerate(SCHULDEN.items()))
    subtotal("Totaal schulden", schulden)
    gap()

    netto = total_bezittingen - schulden
    subtotal("NETTO VERMOGEN", netto, color=C_DARK, text_color="FFFFFF", size=13)


def write_vorderingen(ws):
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 16
    ws.freeze_panes = "A2"

    for c, h in enumerate(["Datum", "Omschrijving", "Bedrag", "Status"], 1):
        cell = ws.cell(1, c, h)
        cell.fill = _fill(C_DARK)
        cell.font = _font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 20

    open_total = 0
    for i, v in enumerate(VORDERINGEN_OP_ODS, 2):
        bg = _fill(C_ALT) if i % 2 == 0 else None
        for c, val in enumerate([v["datum"], v["omschrijving"], v["bedrag"], v.get("status", "Open")], 1):
            cell = ws.cell(i, c, val)
            cell.font = _font()
            if bg:
                cell.fill = bg
            if c == 3:
                cell.number_format = EURO
                cell.alignment = Alignment(horizontal="right")
        if v.get("status", "Open") == "Open":
            open_total += v["bedrag"]

    r_tot = len(VORDERINGEN_OP_ODS) + 2
    ws.cell(r_tot, 2, "Totaal openstaand (ODS schuldig aan Chris)")
    ac = ws.cell(r_tot, 3, open_total)
    for c in range(1, 5):
        cell = ws.cell(r_tot, c)
        cell.font = _font(bold=True)
        cell.fill = _fill(C_SUB)
    ac.number_format = EURO
    ac.alignment = Alignment(horizontal="right")

    ws.auto_filter.ref = "A1:D1"


def main():
    abn_balances              = read_abn_balances()
    revolut_bal, revolut_date = _read_revolut_balance()
    degiro_netto, degiro_date = _read_degiro_netto()
    bitvavo_eur, bitvavo_date = _read_bitvavo_eur_balance()

    print(f"\nNetto Vermogen — {PEILDATUM}")
    print("=" * 50)

    if abn_balances:
        for acc, bal in abn_balances.items():
            print(f"  {ABN_LABELS.get(acc, acc):<38} €{bal:>10,.2f}")
    else:
        print("  ABN AMRO: geen TAB bestanden gevonden")

    if revolut_bal is not None:
        print(f"  {'Revolut':<38} €{revolut_bal:>10,.2f}")
    else:
        print("  Revolut: geen CSV gevonden")

    if ODS_SALDO:
        print(f"  {'Knab Zakelijke Rekening':<38} €{ODS_SALDO:>10,.2f}")
    else:
        print("  Knab: ODS_SALDO = 0 — vul in networth_config.py")

    if degiro_netto is not None:
        print(f"  {'DeGiro (netto kasstroom trades)':<38} €{degiro_netto:>10,.2f}")
    else:
        print("  DeGiro: geen CSV gevonden")

    if bitvavo_eur is not None:
        print(f"  {'Bitvavo (EUR kassaldo)':<38} €{bitvavo_eur:>10,.2f}")
    else:
        print("  Bitvavo: geen CSV gevonden")

    wb = Workbook()
    write_vermogen(
        wb.active, abn_balances,
        revolut_bal=revolut_bal, revolut_date=revolut_date,
        degiro_netto=degiro_netto, degiro_date=degiro_date,
        bitvavo_eur=bitvavo_eur, bitvavo_date=bitvavo_date,
    )
    wb.active.title = "Vermogen"

    write_vorderingen(wb.create_sheet("Vorderingen op ODS"))

    out = OUTPUT_DIR / f"networth_{PEILDATUM}.xlsx"
    wb.save(out)

    open_total = sum(v["bedrag"] for v in VORDERINGEN_OP_ODS if v.get("status", "Open") == "Open")
    print(f"\n  Openstaande vordering op ODS: €{open_total:.2f}")
    print(f"  Opgeslagen: {out}\n")


if __name__ == "__main__":
    main()
