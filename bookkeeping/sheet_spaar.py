from datetime import datetime

from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

from config import format_period, _fill
from excel_output import KLEUR_POSITIEF, KLEUR_NEGATIEF


def write_spaar_overview_sheet(ws, df, year_label="", tx_sheet_name=""):
    ws.title = f"Maand Overzicht Spaar{' ' + year_label if year_label else ''}"
    ws.freeze_panes = "B2"

    if year_label:
        periods = [f"{year_label}-{m:02d}" for m in range(1, 13)]
    else:
        periods = sorted(df["month"].unique()) if len(df) > 0 else []

    now = datetime.now()
    past_count = sum(
        1 for p in periods
        if not ((int(p[:4]) > now.year) or (int(p[:4]) == now.year and int(p[5:]) > now.month))
    )

    n        = len(periods)
    tot_col  = n + 2
    gem_col  = n + 3
    euro_fmt = '€ #,##0;-€ #,##0;""'
    tx       = f"'{tx_sheet_name}'" if ' ' in tx_sheet_name else tx_sheet_name

    def _period_dates(p):
        yr, mo = int(p[:4]), int(p[5:])
        return yr, mo, yr if mo < 12 else yr + 1, mo + 1 if mo < 12 else 1

    def _sumifs(cat, p):
        yr, mo, nyr, nmo = _period_dates(p)
        return (f'=SUMIFS({tx}!$E:$E,'
                f'{tx}!$F:$F,"{cat}",'
                f'{tx}!$A:$A,">="&DATE({yr},{mo},1),'
                f'{tx}!$A:$A,"<"&DATE({nyr},{nmo},1))')

    cats = sorted(df["category"].unique()) if len(df) > 0 else []

    r = 1
    ws.cell(row=r, column=1, value="Categorie").font = Font(bold=True, name="Arial", size=10)
    for c, p in enumerate(periods, 2):
        cell = ws.cell(row=r, column=c, value=format_period(p, "month"))
        cell.font = Font(bold=True, name="Arial", size=10)
        cell.alignment = Alignment(horizontal="center")
    ws.cell(row=r, column=tot_col, value="Totaal").font  = Font(bold=True, name="Arial", size=10)
    ws.cell(row=r, column=tot_col).alignment = Alignment(horizontal="center")
    ws.cell(row=r, column=gem_col, value="Gemiddeld").font = Font(bold=True, name="Arial", size=10)
    ws.cell(row=r, column=gem_col).alignment = Alignment(horizontal="center")
    ws.row_dimensions[r].height = 20

    cat_rows = {}
    for cat in cats:
        r += 1
        cat_rows[cat] = r
        net  = float(df[df["category"] == cat]["amount"].sum())
        fill = _fill(KLEUR_POSITIEF) if net >= 0 else _fill(KLEUR_NEGATIEF)
        lbl  = ws.cell(row=r, column=1, value=cat)
        lbl.font = Font(name="Arial", size=9)
        lbl.fill = fill
        for ci, p in enumerate(periods, 2):
            cell = ws.cell(row=r, column=ci, value=_sumifs(cat, p))
            cell.number_format = euro_fmt
            cell.font = Font(name="Arial", size=9)
            cell.alignment = Alignment(horizontal="right")
        fc = get_column_letter(2)
        lc = get_column_letter(n + 1)
        tc = ws.cell(row=r, column=tot_col, value=f"=SUM({fc}{r}:{lc}{r})")
        tc.number_format = euro_fmt
        tc.font = Font(bold=True, name="Arial", size=9)
        tc.alignment = Alignment(horizontal="right")
        denom = past_count if past_count > 0 else 1
        gc = ws.cell(row=r, column=gem_col, value=f"={get_column_letter(tot_col)}{r}/{denom}")
        gc.number_format = euro_fmt
        gc.font = Font(italic=True, name="Arial", size=9)
        gc.alignment = Alignment(horizontal="right")

    if cat_rows:
        r += 1
        font = Font(bold=True, name="Arial", size=9)
        ws.cell(row=r, column=1, value="Totaal").font = font
        for ci in range(2, n + 2):
            cl    = get_column_letter(ci)
            parts = [f"{cl}{rr}" for rr in cat_rows.values()]
            cell  = ws.cell(row=r, column=ci, value=f"=SUM({','.join(parts)})")
            cell.number_format = euro_fmt
            cell.font = font
            cell.alignment = Alignment(horizontal="right")
        fc = get_column_letter(2)
        lc = get_column_letter(n + 1)
        tc = ws.cell(row=r, column=tot_col, value=f"=SUM({fc}{r}:{lc}{r})")
        tc.number_format = euro_fmt
        tc.font = font
        denom = past_count if past_count > 0 else 1
        gc = ws.cell(row=r, column=gem_col, value=f"={get_column_letter(tot_col)}{r}/{denom}")
        gc.number_format = euro_fmt
        gc.font = font

    ws.column_dimensions["A"].width = 26
    for c in range(2, gem_col + 1):
        ws.column_dimensions[get_column_letter(c)].width = 12
