from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

from config import C_HEADER, WHITE, C_INC_ROW, C_ALT_ROW, ACCOUNT_LABELS, _fill


def write_transactions_sheet(ws, df):
    ws.title = "Transacties"
    ws.freeze_panes = "A2"

    headers = ["Datum", "Rekening", "Omschrijving", "Merchant",
               "Bedrag (EUR)", "Categorie", "Bron", "Maand"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = _fill(C_HEADER)
        cell.font = Font(bold=True, color=WHITE, name="Arial", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 20

    inc_fill = _fill(C_INC_ROW)
    alt_fill = _fill(C_ALT_ROW)
    euro_fmt = '€ #,##0.00;-€ #,##0.00'

    for r, (_, row) in enumerate(df.iterrows(), 2):
        is_income = row["amount"] > 0
        fill = (inc_fill if is_income
                else (alt_fill if r % 2 == 0 else None))
        data = [
            row["date"].date(),
            ACCOUNT_LABELS.get(str(row["account"]), str(row["account"])),
            row["description"][:80],
            row["merchant"], row["amount"], row["category"],
            row.get("source", "ABN"),
            row["month"],
        ]
        for c, val in enumerate(data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = Font(name="Arial", size=9)
            cell.alignment = Alignment(vertical="center")
            if fill:
                cell.fill = fill
            if c == 5:
                cell.number_format = euro_fmt

    for c, w in enumerate([12, 12, 50, 28, 14, 22, 8, 10], 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
