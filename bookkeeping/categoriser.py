import os
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font

from config import RULES_FILE, OWN_ACCOUNTS, C_HEADER, WHITE, _fill, get_iban_rules


def load_rules():
    if not os.path.exists(RULES_FILE):
        print(f"Geen rules.xlsx gevonden op {RULES_FILE}")
        print("Voer uit met --create-rules om een startbestand te maken.")
        return []

    df = pd.read_excel(RULES_FILE, sheet_name="Rules")
    rules = []
    for _, row in df.iterrows():
        kw  = str(row.get("keyword",  "")).strip()
        cat = str(row.get("category", "")).strip()
        if kw and cat:
            rules.append((kw.lower(), cat))
    print(f"Geladen: {len(rules)} categorisatieregels")
    return rules


def categorise(merchant, rules):
    ml = merchant.lower()
    for kw, cat in rules:
        if kw in ml:
            return cat
    return "Dagelijks Overig"


def _categorise_by_desc(description, iban_rules):
    dl = str(description).lower()
    for kw, cat in iban_rules:
        if kw.lower() in dl:
            return cat
    return None


def apply_categories(df, rules):
    iban_rules = get_iban_rules()

    def _cat(row):
        cat = _categorise_by_desc(row["description"], iban_rules)
        return cat if cat else categorise(row["merchant"], rules)

    df["category"] = df.apply(_cat, axis=1)
    # ZZP Opname is only valid for incoming transactions
    df.loc[(df["category"] == "ZZP Opname") & (df["amount"] < 0), "category"] = "Dagelijks Overig"
    print(f"Gecategoriseerd: {len(df)} transacties  (fallback → Dagelijks Overig)")
    return df


def detect_internal_transfers(df):
    mask = pd.Series(False, index=df.index)
    for (date_val, abs_amt), group in df.groupby(
        [df["date"].dt.date, df["amount"].abs()]
    ):
        if len(group) >= 2:
            accs = set(group["account"].unique())
            if len(accs.intersection(OWN_ACCOUNTS)) >= 2:
                if abs(group["amount"].sum()) < 0.02:
                    mask.loc[group.index] = True
    n = int(mask.sum())
    if n:
        df.loc[mask, "category"] = "Interne Overboeking"
        print(f"Interne overboekingen: {n} transacties "
              f"(betaalrekening <-> spaarrekening, uitgesloten van netto)")
    return df


CATEGORY_MIGRATION = {
    "inkomen":                   "Salaris",
    "inkomsten":                 "Salaris",
    "inkomsten overig":          "Inkomsten Overig",
    "verzekeringen":             "Zorgverzekering",
    "telefoon":                  "Telefoon & Internet",
    "kamerhuur":                 "Huur",
    "huur & wonen":              "Huur",
    "vaste lasten":              "Abonnementen",
    "belastingen":               "Dagelijks Overig",
    "sport":                     "Sport & Fitness",
    "ov & reizen":               "OV & Reizen",
    "gezondheid":                "Gezondheid",
    "boodschappen":              "Boodschappen",
    "eten & drinken":            "Eten & Drinken",
    "bankkosten":                "Bankkosten",
    "diversen":                  "Dagelijks Overig",
    "sparen":                    "Sparen",
    "online winkelen":           "Dagelijks Overig",
    "tabak":                     "Eten & Drinken",
    "cultuur & entertainment":   "Cultuur & Entertainment",
}

KEYWORD_OVERRIDES = {
    "tabakshop":         "Eten & Drinken",
    "gogo tabak":        "Eten & Drinken",
    "bol.com":           "Dagelijks Overig",
    "media markt":       "Dagelijks Overig",
    "laurenskerk":       "Cultuur & Entertainment",
    "belastingdienst":   "Dagelijks Overig",
    "sportcity":         "Sport & Fitness",
    "david lloyd":       "Sport & Fitness",
    "youfone":           "Telefoon & Internet",
    "unive":             "Zorgverzekering",
    "huiver":            "Huur",
    "hogeschool":        "Salaris",
    "salaris":           "Salaris",
    "abn amro":          "Bankkosten",
}


def create_starter_rules():
    starter = [
        # ── INKOMEN ──────────────────────────────────────────────────────────
        ("salaris",                 "Salaris"),
        ("hogeschool",              "Salaris"),
        ("loon",                    "Salaris"),
        ("duo",                     "DUO / Studiefinanciering"),
        ("studiefinanciering",      "DUO / Studiefinanciering"),
        ("zorgtoeslag",             "Zorgtoeslag"),
        ("toeslagen",               "Zorgtoeslag"),
        # ── VASTE LASTEN ─────────────────────────────────────────────────────
        ("huiver",                  "Huur"),
        ("kamer",                   "Huur"),
        ("unive",                   "Zorgverzekering"),
        ("cz ",                     "Zorgverzekering"),
        ("zilveren kruis",          "Zorgverzekering"),
        ("youfone",                 "Telefoon & Internet"),
        ("t-mobile",                "Telefoon & Internet"),
        ("abn amro",                "Bankkosten"),
        ("spotify",                 "Abonnementen"),
        ("netflix",                 "Abonnementen"),
        ("ziggo",                   "Abonnementen"),
        ("sportcity",               "Sport & Fitness"),
        ("david lloyd",             "Sport & Fitness"),
        # ── DAGELIJKS ────────────────────────────────────────────────────────
        ("albert heijn",            "Boodschappen"),
        ("dirk",                    "Boodschappen"),
        ("jumbo",                   "Boodschappen"),
        ("lidl",                    "Boodschappen"),
        ("ah ",                     "Boodschappen"),
        ("tuda fruta",              "Boodschappen"),
        ("takeaway",                "Eten & Drinken"),
        ("deliveroo",               "Eten & Drinken"),
        ("ming kee",                "Eten & Drinken"),
        ("cn bagijn",               "Eten & Drinken"),
        ("koffiehuis",              "Eten & Drinken"),
        ("crow-bar",                "Eten & Drinken"),
        ("ls nine bar",             "Eten & Drinken"),
        ("cafe",                    "Eten & Drinken"),
        ("tabakshop",               "Eten & Drinken"),
        ("gogo tabak",              "Eten & Drinken"),
        ("ns groep",                "OV & Reizen"),
        ("ns reizigers",            "OV & Reizen"),
        ("bck*ns",                  "OV & Reizen"),
        ("den haag cs",             "OV & Reizen"),
        ("bol.com",                 "Dagelijks Overig"),
        ("media markt",             "Dagelijks Overig"),
        ("heilzaam",                "Gezondheid"),
        ("apotheek",                "Gezondheid"),
        ("laurenskerk",             "Cultuur & Entertainment"),
        # ── OVERIG ───────────────────────────────────────────────────────────
        ("sparen",                  "Sparen"),
        ("belastingdienst",         "Dagelijks Overig"),
        ("primera",                 "Dagelijks Overig"),
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "Rules"
    ws.cell(row=1, column=1, value="keyword")
    ws.cell(row=1, column=2, value="category")
    for c in [1, 2]:
        ws.cell(row=1, column=c).fill = _fill(C_HEADER)
        ws.cell(row=1, column=c).font = Font(bold=True, color=WHITE, name="Arial")

    for r, (kw, cat) in enumerate(starter, 2):
        ws.cell(row=r, column=1, value=kw)
        ws.cell(row=r, column=2, value=cat)

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 30

    ws2 = wb.create_sheet("Instructies")
    instructions = [
        ("Hoe werkt rules.xlsx?", ""),
        ("", ""),
        ("keyword",  "De tekst in de merchant naam (niet hoofdlettergevoelig, deelwoord)"),
        ("category", "De post die je aan die transactie wilt geven"),
        ("", ""),
        ("Tips:", ""),
        ("Volgorde telt",      "De EERSTE match wint — zet specifieke regels bovenaan"),
        ("Inkomsten",          "Gebruik 'Salaris', 'DUO / Studiefinanciering' etc."),
        ("Nieuw toevoegen",    "Voeg een rij toe aan het Rules tabblad en herrun"),
    ]
    for r, (a, b) in enumerate(instructions, 1):
        ws2.cell(row=r, column=1, value=a).font = Font(bold=(r == 1), name="Arial")
        ws2.cell(row=r, column=2, value=b).font = Font(name="Arial")
    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 65
    wb.save(RULES_FILE)
    print(f"Starter rules.xlsx aangemaakt: {RULES_FILE}")


def migrate_rules():
    if not os.path.exists(RULES_FILE):
        print("Geen rules.xlsx gevonden.")
        return

    wb = load_workbook(RULES_FILE)
    if "Rules" not in wb.sheetnames:
        print("Geen 'Rules' tabblad gevonden in rules.xlsx.")
        return

    ws      = wb["Rules"]
    changed = 0
    for row in ws.iter_rows(min_row=2):
        kw_cell  = row[0]
        cat_cell = row[1]
        kw  = str(kw_cell.value  or "").strip().lower()
        old = str(cat_cell.value or "").strip()

        if kw in KEYWORD_OVERRIDES:
            new = KEYWORD_OVERRIDES[kw]
        else:
            new = CATEGORY_MIGRATION.get(old.lower(), old)

        if new != old:
            cat_cell.value = new
            changed += 1

    wb.save(RULES_FILE)
    print(f"rules.xlsx bijgewerkt: {changed} categorie(en) hernoemd")
