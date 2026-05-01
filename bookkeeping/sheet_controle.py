from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

from config import (
    INCOME_CATS, VASTE_LASTEN_CATS, DAGELIJKS_CATS, OVERIG_CATS,
    WHITE, C_HEADER, format_period, _fill,
)


def write_controle_sheet(ws, df, year_label=""):
    ws.title = f"Controle{' ' + year_label if year_label else ''}"
    ws.freeze_panes = "B2"
    euro_fmt = '€ #,##0;-€ #,##0'

    periods = sorted(df["month"].unique())
    df_real = df[df["category"] != "Interne Overboeking"]
    df_int  = df[df["category"] == "Interne Overboeking"]

    def _s(frame, period):
        return float(frame[frame["month"] == period]["amount"].sum())

    headers = [
        "Maand", "Inkomen (cat)", "Uitgaven (cat)", "Netto Alle", "Interne OB",
    ]

    r = 1
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=r, column=c, value=h)
        cell.fill = _fill(C_HEADER)
        cell.font = Font(bold=True, color=WHITE, name="Arial", size=9)
        cell.alignment = Alignment(horizontal="center" if c > 1 else "left")
    ws.row_dimensions[r].height = 18

    EXPENSE_CATS = VASTE_LASTEN_CATS + DAGELIJKS_CATS + OVERIG_CATS
    running = [0.0] * (len(headers) - 1)

    for period in periods:
        r += 1
        inc_cat = _s(df_real[df_real["category"].isin(INCOME_CATS)], period)
        uit_cat = _s(df_real[df_real["category"].isin(EXPENSE_CATS)], period)
        netto   = _s(df_real, period)
        interne = _s(df_int, period)

        vals = [inc_cat, uit_cat, netto, interne]
        ws.cell(row=r, column=1, value=format_period(period, "month")).font = Font(name="Arial", size=9)
        for c, v in enumerate(vals, 2):
            cell = ws.cell(row=r, column=c, value=round(v))
            cell.number_format = euro_fmt
            cell.font = Font(name="Arial", size=9)
            cell.alignment = Alignment(horizontal="right")
        for i, v in enumerate(vals):
            running[i] += v

    r += 1
    ws.cell(row=r, column=1, value="Totaal").font = Font(bold=True, name="Arial", size=9)
    for c, v in enumerate(running, 2):
        cell = ws.cell(row=r, column=c, value=round(v))
        cell.number_format = euro_fmt
        cell.font = Font(bold=True, name="Arial", size=9)
        cell.alignment = Alignment(horizontal="right")

    ws.column_dimensions["A"].width = 12
    for c in range(2, len(headers) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 15
