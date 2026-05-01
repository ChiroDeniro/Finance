"""
Extra Excel sheets voor de Knab zakelijke rekening:
  - Jaar Vergelijking : year-over-year comparison
  - Uitschieters      : all transactions ranked by size, outliers flagged
"""

from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from config import (
    INCOME_CATS, VASTE_LASTEN_CATS, DAGELIJKS_CATS, OVERIG_CATS,
    C_INC_HDR, C_INC_ROW, C_VL_HDR, C_VL_ROW,
    C_DAG_HDR, C_DAG_ROW, C_OVR_HDR, C_OVR_ROW,
    C_IO_HDR, C_IO_ROW, C_SAM_HDR, C_NETTO_POS, C_NETTO_NEG,
    C_HEADER, WHITE, _fill,
)

EURO_FMT = '€ #,##0.00;[Red]-€ #,##0.00'
PCT_FMT  = '+0%;-0%'
DATE_FMT = 'DD-MM-YYYY'


def _hdr(ws, row, col, value, fill=C_HEADER, align="left"):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(bold=True, color=WHITE, name="Arial", size=9)
    cell.fill = _fill(fill)
    cell.alignment = Alignment(horizontal=align)
    return cell


def _cell(ws, row, col, value=None, bold=False, fill=None, fmt=None, align="right"):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(bold=bold, name="Arial", size=9)
    cell.alignment = Alignment(horizontal=align)
    if fill:
        cell.fill = _fill(fill)
    if fmt:
        cell.number_format = fmt
    return cell


def write_knab_jaaroverzicht(ws, df):
    """Year-over-year income/expense comparison for the Knab business account."""
    ws.freeze_panes = "B2"

    years = sorted(df["year"].unique())
    n_yr  = len(years)

    # ── Header row ────────────────────────────────────────────────────────────
    _hdr(ws, 1, 1, "Categorie")
    for c, yr in enumerate(years, 2):
        _hdr(ws, 1, c, yr, align="right")
    for i in range(n_yr - 1):
        _hdr(ws, 1, n_yr + 1 + i, f"Δ {years[i]}→{years[i+1]}", align="right")
    ws.row_dimensions[1].height = 18

    df_real = df[df["category"] != "Interne Overboeking"]

    groups = [
        ("INKOMSTEN",         INCOME_CATS,       C_INC_HDR, C_INC_ROW),
        ("VASTE LASTEN",      VASTE_LASTEN_CATS, C_VL_HDR,  C_VL_ROW),
        ("DAGELIJKS",         DAGELIJKS_CATS,    C_DAG_HDR, C_DAG_ROW),
        ("OVERIG / ZAKELIJK", OVERIG_CATS,       C_OVR_HDR, C_OVR_ROW),
    ]

    row = 2
    for grp_name, cats, hdr_color, row_color in groups:
        cats_with_data = [c for c in cats if (df_real["category"] == c).any()]
        if not cats_with_data:
            continue

        # Section header spanning all columns
        for c in range(1, n_yr + 1 + (n_yr - 1) + 1):
            _hdr(ws, row, c, grp_name if c == 1 else "", fill=hdr_color)
        row += 1

        for cat in cats_with_data:
            _cell(ws, row, 1, cat, align="left").fill = _fill(row_color)
            totals = []
            for c, yr in enumerate(years, 2):
                total = df_real[(df_real["category"] == cat) & (df_real["year"] == yr)]["amount"].sum()
                totals.append(total)
                cell = _cell(ws, row, c, total, fmt=EURO_FMT, fill=row_color)
            # Growth columns
            for i in range(len(totals) - 1):
                c = n_yr + 1 + i
                if totals[i] != 0:
                    pct = (totals[i + 1] - totals[i]) / abs(totals[i])
                    _cell(ws, row, c, pct, fmt=PCT_FMT, fill=row_color)
                else:
                    _cell(ws, row, c, fill=row_color)
            row += 1

    # ── Interne Overboekingen ─────────────────────────────────────────────────
    for c in range(1, n_yr + 1 + (n_yr - 1) + 1):
        _hdr(ws, row, c, "INTERNE OVERBOEKINGEN" if c == 1 else "", fill=C_IO_HDR)
    row += 1
    _cell(ws, row, 1, "Interne Overboeking", align="left").fill = _fill(C_IO_ROW)
    io_totals = []
    for c, yr in enumerate(years, 2):
        total = df[(df["category"] == "Interne Overboeking") & (df["year"] == yr)]["amount"].sum()
        io_totals.append(total)
        _cell(ws, row, c, total, fmt=EURO_FMT, fill=C_IO_ROW)
    row += 1

    # ── NETTO ─────────────────────────────────────────────────────────────────
    row += 1
    for c in range(1, n_yr + 1 + (n_yr - 1) + 1):
        _hdr(ws, row, c, "NETTO" if c == 1 else "", fill=C_SAM_HDR)

    nettos = []
    for c, yr in enumerate(years, 2):
        netto = df_real[df_real["year"] == yr]["amount"].sum()
        nettos.append(netto)
        color = C_NETTO_POS if netto >= 0 else C_NETTO_NEG
        cell = ws.cell(row=row, column=c, value=netto)
        cell.font = Font(bold=True, color=WHITE, name="Arial", size=9)
        cell.fill = _fill(color)
        cell.number_format = EURO_FMT
        cell.alignment = Alignment(horizontal="right")

    for i in range(len(nettos) - 1):
        c = n_yr + 1 + i
        if nettos[i] != 0:
            pct = (nettos[i + 1] - nettos[i]) / abs(nettos[i])
            color = C_NETTO_POS if nettos[i + 1] >= nettos[i] else C_NETTO_NEG
            cell = ws.cell(row=row, column=c, value=pct)
            cell.font = Font(bold=True, color=WHITE, name="Arial", size=9)
            cell.fill = _fill(color)
            cell.number_format = PCT_FMT
            cell.alignment = Alignment(horizontal="right")

    # ── Column widths ─────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 26
    for c in range(2, n_yr + 1 + n_yr + 1):
        ws.column_dimensions[get_column_letter(c)].width = 16


def write_knab_uitschieters(ws, df):
    """All transactions ranked by absolute amount; outliers flagged (>1.5σ above mean)."""
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False

    df_real = df[df["category"] != "Interne Overboeking"].copy()
    df_real["abs_amount"] = df_real["amount"].abs()

    # Outlier threshold: mean + 1.5 × std of absolute amounts
    if len(df_real) >= 4:
        threshold = df_real["abs_amount"].mean() + 1.5 * df_real["abs_amount"].std()
    else:
        threshold = float("inf")

    headers = ["Datum", "Merchant / Tegenpartij", "Categorie", "Bedrag (EUR)", "Opmerking"]
    col_align = ["left", "left", "left", "right", "left"]
    for c, (h, a) in enumerate(zip(headers, col_align), 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = _fill(C_HEADER)
        cell.font = Font(bold=True, color=WHITE, name="Arial", size=9)
        cell.alignment = Alignment(horizontal=a)
    ws.row_dimensions[1].height = 18

    df_sorted = df_real.sort_values("abs_amount", ascending=False)

    for r, (_, tx) in enumerate(df_sorted.iterrows(), 2):
        is_income    = tx["amount"] > 0
        is_outlier   = tx["abs_amount"] >= threshold
        fill_color   = "E2EFDA" if is_income else ("FCE4D6" if is_outlier else ("EEF3FA" if r % 2 == 0 else "FFFFFF"))
        opmerking    = "★ Uitschieter" if is_outlier else ""

        datum = tx["date"]
        if hasattr(datum, "date"):
            datum = datum.date()

        ws.cell(row=r, column=1, value=datum).number_format = DATE_FMT
        ws.cell(row=r, column=2, value=str(tx["merchant"])[:60])
        ws.cell(row=r, column=3, value=tx["category"])
        ws.cell(row=r, column=4, value=float(tx["amount"])).number_format = EURO_FMT
        ws.cell(row=r, column=5, value=opmerking)

        for c in range(1, 6):
            cell = ws.cell(row=r, column=c)
            cell.fill = _fill(fill_color)
            cell.font = Font(name="Arial", size=9)
            cell.alignment = Alignment(
                horizontal="right" if c == 4 else "left"
            )

    ws.column_dimensions["A"].width = 13
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 16
