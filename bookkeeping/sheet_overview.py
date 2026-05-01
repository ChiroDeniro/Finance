from datetime import datetime

from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config import (
    INCOME_CATS, VASTE_LASTEN_CATS, DAGELIJKS_CATS, OVERIG_CATS,
    KOSTEN_EXCLUDE, format_period, _fill,
)
from excel_output import KLEUR_POSITIEF, KLEUR_NEGATIEF


def write_overview_sheet(ws, df, group_by="month", tx_sheet_name="Transacties", year_label=""):
    if group_by == "month":
        ws.title = "Maand Overzicht"
        if year_label:
            periods = [f"{year_label}-{m:02d}" for m in range(1, 13)]
        else:
            periods = sorted(df["month"].unique())
    else:
        ws.title  = "Jaar Overzicht"
        periods   = sorted(df["year"].unique())

    ws.freeze_panes = "B2"

    now = datetime.now()

    if group_by == "month":
        def _is_future(p):
            yr, mo = int(p[:4]), int(p[5:])
            return (yr > now.year) or (yr == now.year and mo > now.month)
        past_count = sum(1 for p in periods if not _is_future(p))
    else:
        past_count = len(periods)

    n        = len(periods)
    tot_col  = n + 2
    gem_col  = n + 3 if group_by == "month" else None
    last_col = gem_col or tot_col
    euro_fmt = '€ #,##0;-€ #,##0;""'

    thin_bottom = Border(bottom=Side(style="thin"))

    tx = f"'{tx_sheet_name}'" if ' ' in tx_sheet_name else tx_sheet_name

    def _period_dates(p):
        if group_by == "month":
            yr, mo = int(p[:4]), int(p[5:])
            nyr = yr if mo < 12 else yr + 1
            nmo = mo + 1 if mo < 12 else 1
        else:
            yr, mo, nyr, nmo = int(p), 1, int(p) + 1, 1
        return yr, mo, nyr, nmo

    def _sumifs(cat, p):
        yr, mo, nyr, nmo = _period_dates(p)
        return (f'=SUMIFS({tx}!$E:$E,'
                f'{tx}!$F:$F,"{cat}",'
                f'{tx}!$A:$A,">="&DATE({yr},{mo},1),'
                f'{tx}!$A:$A,"<"&DATE({nyr},{nmo},1))')

    row_num      = [2]
    row_for_cat  = {}
    subtotal_row = {}
    sam_row_map  = {}

    def _header_row():
        r = row_num[0]
        ws.cell(row=r, column=1, value="Categorie")
        ws.cell(row=r, column=1).font = Font(bold=True, name="Arial", size=10)
        for c, p in enumerate(periods, 2):
            cell = ws.cell(row=r, column=c, value=format_period(p, group_by))
            cell.font = Font(bold=True, name="Arial", size=10)
            cell.alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=tot_col, value="Totaal")
        ws.cell(row=r, column=tot_col).font = Font(bold=True, name="Arial", size=10)
        ws.cell(row=r, column=tot_col).alignment = Alignment(horizontal="center")
        if gem_col:
            ws.cell(row=r, column=gem_col, value="Gemiddeld")
            ws.cell(row=r, column=gem_col).font = Font(bold=True, name="Arial", size=10)
            ws.cell(row=r, column=gem_col).alignment = Alignment(horizontal="center")
        ws.row_dimensions[r].height = 20
        row_num[0] += 1

    def _section_hdr(label):
        r = row_num[0]
        for c in range(1, last_col + 1):
            ws.cell(row=r, column=c).font = Font(bold=True, name="Arial", size=9)
            ws.cell(row=r, column=c).border = thin_bottom
        ws.cell(row=r, column=1).value = label
        ws.row_dimensions[r].height = 16
        row_num[0] += 1

    def _data_row(cat):
        r = row_num[0]
        row_for_cat[cat] = r
        ws.cell(row=r, column=1, value=cat)
        ws.cell(row=r, column=1).font = Font(name="Arial", size=9)
        for ci, p in enumerate(periods, 2):
            cell = ws.cell(row=r, column=ci, value=_sumifs(cat, p))
            cell.number_format = euro_fmt
            cell.font = Font(name="Arial", size=9)
            cell.alignment = Alignment(horizontal="right")
        fc = get_column_letter(2)
        lc = get_column_letter(n + 1)
        tc = ws.cell(row=r, column=tot_col, value=f"=SUM({fc}{r}:{lc}{r})")
        tc.number_format = euro_fmt
        tc.font = Font(bold=True, name="Arial", size=9)
        tc.alignment = Alignment(horizontal="right")
        if gem_col:
            denom = past_count if past_count > 0 else 1
            gc = ws.cell(row=r, column=gem_col,
                         value=f"={get_column_letter(tot_col)}{r}/{denom}")
            gc.number_format = euro_fmt
            gc.font = Font(italic=True, name="Arial", size=9)
            gc.alignment = Alignment(horizontal="right")
        row_num[0] += 1

    def _subtotal_row(label, cats, color, sec_key=None):
        r = row_num[0]
        if sec_key:
            subtotal_row[sec_key] = r
        fill = _fill(color) if color else None
        font = Font(bold=True, name="Arial", size=9)
        lbl = ws.cell(row=r, column=1, value=label)
        lbl.font = font
        if fill:
            lbl.fill = fill
        for ci in range(2, n + 2):
            cl = get_column_letter(ci)
            parts = [f"{cl}{row_for_cat[c]}" for c in cats if c in row_for_cat]
            cell = ws.cell(row=r, column=ci,
                           value=f"=SUM({','.join(parts)})" if parts else "=0")
            cell.number_format = euro_fmt
            cell.font = font
            if fill:
                cell.fill = fill
        fc = get_column_letter(2)
        lc = get_column_letter(n + 1)
        tc = ws.cell(row=r, column=tot_col, value=f"=SUM({fc}{r}:{lc}{r})")
        tc.number_format = euro_fmt
        tc.font = font
        if fill:
            tc.fill = fill
        if gem_col:
            denom = past_count if past_count > 0 else 1
            gc = ws.cell(row=r, column=gem_col,
                         value=f"={get_column_letter(tot_col)}{r}/{denom}")
            gc.number_format = euro_fmt
            gc.font = font
            if fill:
                gc.fill = fill
        ws.row_dimensions[r].height = 15
        row_num[0] += 1

    def _blank():
        ws.row_dimensions[row_num[0]].height = 5
        row_num[0] += 1

    def _samenvatting_row(label, sec_key, color):
        r = row_num[0]
        sam_row_map[sec_key] = r
        fill = _fill(color) if color else None
        font = Font(bold=True, name="Arial", size=9)
        lbl  = ws.cell(row=r, column=1, value=label)
        lbl.font = font
        if fill:
            lbl.fill = fill
        src = subtotal_row[sec_key]
        for ci in range(2, last_col + 1):
            cl   = get_column_letter(ci)
            cell = ws.cell(row=r, column=ci, value=f"={cl}{src}")
            cell.number_format = euro_fmt
            cell.font = font
            if fill:
                cell.fill = fill
        row_num[0] += 1

    _header_row()

    _section_hdr("INKOMEN")
    for cat in INCOME_CATS:
        _data_row(cat)
    _subtotal_row("Totaal Inkomen", INCOME_CATS, KLEUR_POSITIEF, sec_key="inkomen")
    _blank()

    _section_hdr("VASTE LASTEN")
    for cat in VASTE_LASTEN_CATS:
        _data_row(cat)
    _subtotal_row("Totaal Vaste Lasten", VASTE_LASTEN_CATS, KLEUR_NEGATIEF, sec_key="vl")
    _blank()

    _section_hdr("DAGELIJKSE UITGAVEN")
    for cat in DAGELIJKS_CATS:
        _data_row(cat)
    _subtotal_row("Totaal Dagelijks", DAGELIJKS_CATS, KLEUR_NEGATIEF, sec_key="dag")
    _blank()

    _section_hdr("OVERIG")
    for cat in OVERIG_CATS:
        _data_row(cat)
    _subtotal_row("Totaal Overig", OVERIG_CATS, None, sec_key="overig")
    _blank()

    _section_hdr("INTERNE OVERBOEKINGEN")
    _data_row("Interne Overboeking")
    _blank()

    _section_hdr("SAMENVATTING")
    _samenvatting_row("Totaal Inkomen",      "inkomen", KLEUR_POSITIEF)
    _blank()
    _samenvatting_row("Totaal Vaste Lasten", "vl",      KLEUR_NEGATIEF)
    _samenvatting_row("Totaal Dagelijks",    "dag",     KLEUR_NEGATIEF)
    _samenvatting_row("Totaal Overig",       "overig",  None)

    EXPENSE_CATS = VASTE_LASTEN_CATS + DAGELIJKS_CATS + [c for c in OVERIG_CATS if c not in KOSTEN_EXCLUDE]
    r    = row_num[0]
    font = Font(bold=True, name="Arial", size=9)
    ws.cell(row=r, column=1, value="Totaal Kosten").font = font
    for ci in range(2, n + 2):
        cl    = get_column_letter(ci)
        parts = [f"{cl}{row_for_cat[c]}" for c in EXPENSE_CATS if c in row_for_cat]
        cell  = ws.cell(row=r, column=ci,
                        value=f"=ABS(SUM({','.join(parts)}))" if parts else "=0")
        cell.number_format = euro_fmt
        cell.font = font
    fc = get_column_letter(2)
    lc = get_column_letter(n + 1)
    tc = ws.cell(row=r, column=tot_col, value=f"=SUM({fc}{r}:{lc}{r})")
    tc.number_format = euro_fmt
    tc.font = font
    if gem_col:
        denom = past_count if past_count > 0 else 1
        gc = ws.cell(row=r, column=gem_col,
                     value=f"={get_column_letter(tot_col)}{r}/{denom}")
        gc.number_format = euro_fmt
        gc.font = font
    ws.row_dimensions[r].height = 15
    row_num[0] += 1
    _blank()

    df_real     = df[df["category"] != "Interne Overboeking"]
    netto_color = KLEUR_POSITIEF if df_real["amount"].sum() >= 0 else KLEUR_NEGATIEF
    netto_fill  = _fill(netto_color)
    netto_font  = Font(bold=True, name="Arial", size=11)

    r = row_num[0]
    ws.cell(row=r, column=1, value="NETTO").fill = netto_fill
    ws.cell(row=r, column=1).font = netto_font
    ws.row_dimensions[r].height = 18
    for ci, _ in enumerate(periods, 2):
        cl    = get_column_letter(ci)
        parts = [f"{cl}{sam_row_map[k]}"
                 for k in ("inkomen", "vl", "dag", "overig") if k in sam_row_map]
        cell  = ws.cell(row=r, column=ci, value=f"=SUM({','.join(parts)})")
        cell.number_format = euro_fmt
        cell.fill = netto_fill
        cell.font = netto_font
    fc = get_column_letter(2)
    lc = get_column_letter(n + 1)
    tc = ws.cell(row=r, column=tot_col, value=f"=SUM({fc}{r}:{lc}{r})")
    tc.number_format = euro_fmt
    tc.fill = netto_fill
    tc.font = netto_font
    if gem_col:
        denom = past_count if past_count > 0 else 1
        gc = ws.cell(row=r, column=gem_col,
                     value=f"={get_column_letter(tot_col)}{r}/{denom}")
        gc.number_format = euro_fmt
        gc.fill = netto_fill
        gc.font = netto_font

    ws.column_dimensions["A"].width = 26
    for c in range(2, last_col + 1):
        ws.column_dimensions[get_column_letter(c)].width = 12
