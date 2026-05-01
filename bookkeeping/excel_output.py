import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

KLEUR_POSITIEF = "C6EFCE"
KLEUR_NEGATIEF = "FCE4D6"

from config import (
    OUTPUT_DIR,
    INCOME_CATS, VASTE_LASTEN_CATS, DAGELIJKS_CATS, OVERIG_CATS,
    WHITE,
    C_INC_HDR, C_INC_SUB, C_INC_ROW,
    C_VL_HDR,  C_VL_SUB,
    C_DAG_HDR, C_DAG_SUB,
    C_OVR_HDR, C_OVR_SUB,
    C_SAM_HDR, C_NETTO_POS, C_NETTO_NEG,
    C_HEADER, C_ALT_ROW, KOSTEN_EXCLUDE,
    ACCOUNT_LABELS,
    format_period, _fill,
)


def write_transactions_sheet(ws, df):
    ws.title = "Transacties"
    ws.freeze_panes = "A2"

    headers = ["Datum", "Rekening", "Omschrijving", "Merchant",
               "Bedrag (EUR)", "Categorie", "Maand"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = _fill(C_HEADER)
        cell.font = Font(bold=True, color=WHITE, name="Arial", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 20

    inc_fill = _fill(C_INC_ROW)
    alt_fill = _fill(C_ALT_ROW)
    euro_fmt = '€ #,##0.00;-€ #,##0.00'

    for r, (_, row) in enumerate(df.iterrows(), 2):
        is_income = row["amount"] > 0
        fill = (inc_fill if is_income
                else (alt_fill if r % 2 == 0 else None))
        data = [
            row["date"].date(),
            ACCOUNT_LABELS.get(str(row["account"]), str(row["account"])),
            row["description"][:80],
            row["merchant"], row["amount"], row["category"], row["month"],
        ]
        for c, val in enumerate(data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = Font(name="Arial", size=9)
            cell.alignment = Alignment(vertical="center")
            if fill:
                cell.fill = fill
            if c == 5:
                cell.number_format = euro_fmt

    for c, w in enumerate([12, 12, 50, 28, 14, 22, 10], 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


def write_overview_sheet(ws, df, group_by="month", tx_sheet_name="Transacties", year_label=""):
    if group_by == "month":
        ws.title = "Maand Overzicht"
        if year_label:
            periods = [f"{year_label}-{m:02d}" for m in range(1, 13)]
        else:
            periods = sorted(df["month"].unique())
    else:
        ws.title  = "Jaar Overzicht"
        periods   = sorted(df["year"].unique())

    ws.freeze_panes = "B2"

    now = datetime.now()

    if group_by == "month":
        def _is_future(p):
            yr, mo = int(p[:4]), int(p[5:])
            return (yr > now.year) or (yr == now.year and mo > now.month)
        past_count = sum(1 for p in periods if not _is_future(p))
    else:
        past_count = len(periods)

    n        = len(periods)
    tot_col  = n + 2
    gem_col  = n + 3 if group_by == "month" else None
    last_col = gem_col or tot_col
    euro_fmt = '€ #,##0;-€ #,##0;""'

    thin_bottom = Border(bottom=Side(style="thin"))

    tx = f"'{tx_sheet_name}'" if ' ' in tx_sheet_name else tx_sheet_name

    def _period_dates(p):
        if group_by == "month":
            yr, mo = int(p[:4]), int(p[5:])
            nyr = yr if mo < 12 else yr + 1
            nmo = mo + 1 if mo < 12 else 1
        else:
            yr, mo, nyr, nmo = int(p), 1, int(p) + 1, 1
        return yr, mo, nyr, nmo

    def _sumifs(cat, p):
        yr, mo, nyr, nmo = _period_dates(p)
        return (f'=SUMIFS({tx}!$E:$E,'
                f'{tx}!$F:$F,"{cat}",'
                f'{tx}!$A:$A,">="&DATE({yr},{mo},1),'
                f'{tx}!$A:$A,"<"&DATE({nyr},{nmo},1))')

    row_num      = [2]
    row_for_cat  = {}
    subtotal_row = {}
    sam_row_map  = {}

    def _header_row():
        r = row_num[0]
        ws.cell(row=r, column=1, value="Categorie")
        ws.cell(row=r, column=1).font = Font(bold=True, name="Arial", size=10)
        for c, p in enumerate(periods, 2):
            cell = ws.cell(row=r, column=c, value=format_period(p, group_by))
            cell.font = Font(bold=True, name="Arial", size=10)
            cell.alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=tot_col, value="Totaal")
        ws.cell(row=r, column=tot_col).font = Font(bold=True, name="Arial", size=10)
        ws.cell(row=r, column=tot_col).alignment = Alignment(horizontal="center")
        if gem_col:
            ws.cell(row=r, column=gem_col, value="Gemiddeld")
            ws.cell(row=r, column=gem_col).font = Font(bold=True, name="Arial", size=10)
            ws.cell(row=r, column=gem_col).alignment = Alignment(horizontal="center")
        ws.row_dimensions[r].height = 20
        row_num[0] += 1

    def _section_hdr(label):
        r = row_num[0]
        for c in range(1, last_col + 1):
            ws.cell(row=r, column=c).font = Font(bold=True, name="Arial", size=9)
            ws.cell(row=r, column=c).border = thin_bottom
        ws.cell(row=r, column=1).value = label
        ws.row_dimensions[r].height = 16
        row_num[0] += 1

    def _data_row(cat):
        r = row_num[0]
        row_for_cat[cat] = r
        ws.cell(row=r, column=1, value=cat)
        ws.cell(row=r, column=1).font = Font(name="Arial", size=9)
        for ci, p in enumerate(periods, 2):
            cell = ws.cell(row=r, column=ci, value=_sumifs(cat, p))
            cell.number_format = euro_fmt
            cell.font = Font(name="Arial", size=9)
            cell.alignment = Alignment(horizontal="right")
        fc = get_column_letter(2)
        lc = get_column_letter(n + 1)
        tc = ws.cell(row=r, column=tot_col, value=f"=SUM({fc}{r}:{lc}{r})")
        tc.number_format = euro_fmt
        tc.font = Font(bold=True, name="Arial", size=9)
        tc.alignment = Alignment(horizontal="right")
        if gem_col:
            denom = past_count if past_count > 0 else 1
            gc = ws.cell(row=r, column=gem_col,
                         value=f"={get_column_letter(tot_col)}{r}/{denom}")
            gc.number_format = euro_fmt
            gc.font = Font(italic=True, name="Arial", size=9)
            gc.alignment = Alignment(horizontal="right")
        row_num[0] += 1

    def _subtotal_row(label, cats, color, sec_key=None):
        r = row_num[0]
        if sec_key:
            subtotal_row[sec_key] = r
        fill = _fill(color) if color else None
        font = Font(bold=True, name="Arial", size=9)
        lbl = ws.cell(row=r, column=1, value=label)
        lbl.font = font
        if fill:
            lbl.fill = fill
        for ci in range(2, n + 2):
            cl = get_column_letter(ci)
            parts = [f"{cl}{row_for_cat[c]}" for c in cats if c in row_for_cat]
            cell = ws.cell(row=r, column=ci,
                           value=f"=SUM({','.join(parts)})" if parts else "=0")
            cell.number_format = euro_fmt
            cell.font = font
            if fill:
                cell.fill = fill
        fc = get_column_letter(2)
        lc = get_column_letter(n + 1)
        tc = ws.cell(row=r, column=tot_col, value=f"=SUM({fc}{r}:{lc}{r})")
        tc.number_format = euro_fmt
        tc.font = font
        if fill:
            tc.fill = fill
        if gem_col:
            denom = past_count if past_count > 0 else 1
            gc = ws.cell(row=r, column=gem_col,
                         value=f"={get_column_letter(tot_col)}{r}/{denom}")
            gc.number_format = euro_fmt
            gc.font = font
            if fill:
                gc.fill = fill
        ws.row_dimensions[r].height = 15
        row_num[0] += 1

    def _blank():
        ws.row_dimensions[row_num[0]].height = 5
        row_num[0] += 1

    def _samenvatting_row(label, sec_key, color):
        r = row_num[0]
        sam_row_map[sec_key] = r
        fill = _fill(color) if color else None
        font = Font(bold=True, name="Arial", size=9)
        lbl  = ws.cell(row=r, column=1, value=label)
        lbl.font = font
        if fill:
            lbl.fill = fill
        src = subtotal_row[sec_key]
        for ci in range(2, last_col + 1):
            cl   = get_column_letter(ci)
            cell = ws.cell(row=r, column=ci, value=f"={cl}{src}")
            cell.number_format = euro_fmt
            cell.font = font
            if fill:
                cell.fill = fill
        row_num[0] += 1

    _header_row()

    _section_hdr("INKOMEN")
    for cat in INCOME_CATS:
        _data_row(cat)
    _subtotal_row("Totaal Inkomen", INCOME_CATS, KLEUR_POSITIEF, sec_key="inkomen")
    _blank()

    _section_hdr("VASTE LASTEN")
    for cat in VASTE_LASTEN_CATS:
        _data_row(cat)
    _subtotal_row("Totaal Vaste Lasten", VASTE_LASTEN_CATS, KLEUR_NEGATIEF, sec_key="vl")
    _blank()

    _section_hdr("DAGELIJKSE UITGAVEN")
    for cat in DAGELIJKS_CATS:
        _data_row(cat)
    _subtotal_row("Totaal Dagelijks", DAGELIJKS_CATS, KLEUR_NEGATIEF, sec_key="dag")
    _blank()

    _section_hdr("OVERIG")
    for cat in OVERIG_CATS:
        _data_row(cat)
    _subtotal_row("Totaal Overig", OVERIG_CATS, None, sec_key="overig")
    _blank()

    _section_hdr("INTERNE OVERBOEKINGEN")
    _data_row("Interne Overboeking")
    _blank()

    _section_hdr("SAMENVATTING")
    _samenvatting_row("Totaal Inkomen",      "inkomen", KLEUR_POSITIEF)
    _blank()
    _samenvatting_row("Totaal Vaste Lasten", "vl",      KLEUR_NEGATIEF)
    _samenvatting_row("Totaal Dagelijks",    "dag",     KLEUR_NEGATIEF)
    _samenvatting_row("Totaal Overig",       "overig",  None)

    EXPENSE_CATS = VASTE_LASTEN_CATS + DAGELIJKS_CATS + [c for c in OVERIG_CATS if c not in KOSTEN_EXCLUDE]
    r    = row_num[0]
    font = Font(bold=True, name="Arial", size=9)
    ws.cell(row=r, column=1, value="Totaal Kosten").font = font
    for ci in range(2, n + 2):
        cl    = get_column_letter(ci)
        parts = [f"{cl}{row_for_cat[c]}" for c in EXPENSE_CATS if c in row_for_cat]
        cell  = ws.cell(row=r, column=ci,
                        value=f"=ABS(SUM({','.join(parts)}))" if parts else "=0")
        cell.number_format = euro_fmt
        cell.font = font
    fc = get_column_letter(2)
    lc = get_column_letter(n + 1)
    tc = ws.cell(row=r, column=tot_col, value=f"=SUM({fc}{r}:{lc}{r})")
    tc.number_format = euro_fmt
    tc.font = font
    if gem_col:
        denom = past_count if past_count > 0 else 1
        gc = ws.cell(row=r, column=gem_col,
                     value=f"={get_column_letter(tot_col)}{r}/{denom}")
        gc.number_format = euro_fmt
        gc.font = font
    ws.row_dimensions[r].height = 15
    row_num[0] += 1
    _blank()

    df_real     = df[df["category"] != "Interne Overboeking"]
    netto_color = KLEUR_POSITIEF if df_real["amount"].sum() >= 0 else KLEUR_NEGATIEF
    netto_fill  = _fill(netto_color)
    netto_font  = Font(bold=True, name="Arial", size=11)

    r = row_num[0]
    ws.cell(row=r, column=1, value="NETTO").fill = netto_fill
    ws.cell(row=r, column=1).font = netto_font
    ws.row_dimensions[r].height = 18
    for ci, _ in enumerate(periods, 2):
        cl    = get_column_letter(ci)
        parts = [f"{cl}{sam_row_map[k]}"
                 for k in ("inkomen", "vl", "dag", "overig") if k in sam_row_map]
        cell  = ws.cell(row=r, column=ci, value=f"=SUM({','.join(parts)})")
        cell.number_format = euro_fmt
        cell.fill = netto_fill
        cell.font = netto_font
    fc = get_column_letter(2)
    lc = get_column_letter(n + 1)
    tc = ws.cell(row=r, column=tot_col, value=f"=SUM({fc}{r}:{lc}{r})")
    tc.number_format = euro_fmt
    tc.fill = netto_fill
    tc.font = netto_font
    if gem_col:
        denom = past_count if past_count > 0 else 1
        gc = ws.cell(row=r, column=gem_col,
                     value=f"={get_column_letter(tot_col)}{r}/{denom}")
        gc.number_format = euro_fmt
        gc.fill = netto_fill
        gc.font = netto_font

    ws.column_dimensions["A"].width = 26
    for c in range(2, last_col + 1):
        ws.column_dimensions[get_column_letter(c)].width = 12


def write_jaar_samenvatting_sheet(ws, df, year_label="", tx_sheet_name="Transacties"):
    ws.title = f"Jaar Samenvatting{' ' + year_label if year_label else ''}"
    ws.freeze_panes = "B2"

    euro_fmt = '€ #,##0;-€ #,##0'
    pct_fmt  = '0.0"%"'

    EXCLUDE = {"Interne Overboeking"}
    df_real = df[~df["category"].isin(EXCLUDE)]
    cat_totals = df_real.groupby("category")["amount"].sum()

    def _total(cats):
        return sum(cat_totals.get(c, 0.0) for c in cats)

    totaal_inkomen = _total(INCOME_CATS)

    def _pct(val):
        return round(val / totaal_inkomen * 100, 1) if totaal_inkomen else 0.0

    tx = f"'{tx_sheet_name}'" if ' ' in tx_sheet_name else tx_sheet_name

    def _cat_formula(cat):
        if year_label:
            yr = int(year_label)
            return (f'=SUMIFS({tx}!$E:$E,{tx}!$F:$F,"{cat}",'
                    f'{tx}!$A:$A,">="&DATE({yr},1,1),'
                    f'{tx}!$A:$A,"<"&DATE({yr + 1},1,1))')
        return f'=SUMIF({tx}!$F:$F,"{cat}",{tx}!$E:$E)'

    row_num      = [2]
    row_for_cat  = {}
    subtotal_row = {}
    sam_row_map  = {}

    def _hdr():
        r = row_num[0]
        for c, h in enumerate(["Categorie", "Totaal", "% van Inkomen"], 1):
            cell = ws.cell(row=r, column=c, value=h)
            cell.fill = _fill(C_HEADER)
            cell.font = Font(bold=True, color=WHITE, name="Arial", size=10)
            cell.alignment = Alignment(horizontal="center" if c > 1 else "left")
        ws.row_dimensions[r].height = 20
        row_num[0] += 1

    def _sec_hdr(label, color):
        r = row_num[0]
        for c in range(1, 4):
            ws.cell(row=r, column=c).fill = _fill(color)
            ws.cell(row=r, column=c).font = Font(bold=True, color=WHITE, name="Arial", size=9)
        ws.cell(row=r, column=1).value = label
        ws.row_dimensions[r].height = 16
        row_num[0] += 1

    def _data_row(cat, _unused=None):
        r = row_num[0]
        row_for_cat[cat] = r
        ws.cell(row=r, column=1, value=cat)
        ws.cell(row=r, column=1).font = Font(name="Arial", size=9)
        tc = ws.cell(row=r, column=2, value=_cat_formula(cat))
        tc.number_format = euro_fmt
        tc.font = Font(name="Arial", size=9)
        tc.alignment = Alignment(horizontal="right")
        pc = ws.cell(row=r, column=3, value=_pct(cat_totals.get(cat, 0.0)))
        pc.number_format = pct_fmt
        pc.font = Font(italic=True, name="Arial", size=9)
        pc.alignment = Alignment(horizontal="right")
        row_num[0] += 1

    def _subtotal(label, cats, color, sec_key=None):
        r = row_num[0]
        if sec_key:
            subtotal_row[sec_key] = r
        fill = _fill(color)
        font = Font(bold=True, color=WHITE, name="Arial", size=9)
        ws.cell(row=r, column=1, value=label).fill = fill
        ws.cell(row=r, column=1).font = font
        parts = [f"B{row_for_cat[c]}" for c in cats if c in row_for_cat]
        tc = ws.cell(row=r, column=2,
                     value=f"=SUM({','.join(parts)})" if parts else "=0")
        tc.number_format = euro_fmt
        tc.fill = fill
        tc.font = font
        pc = ws.cell(row=r, column=3, value=_pct(_total(cats)))
        pc.number_format = pct_fmt
        pc.fill = fill
        pc.font = font
        ws.row_dimensions[r].height = 15
        row_num[0] += 1

    def _blank():
        ws.row_dimensions[row_num[0]].height = 5
        row_num[0] += 1

    def _sam_row(label, sec_key, sub_color):
        r = row_num[0]
        sam_row_map[sec_key] = r
        fill = _fill(sub_color)
        font = Font(bold=True, color=WHITE, name="Arial", size=9)
        ws.cell(row=r, column=1, value=label).fill = fill
        ws.cell(row=r, column=1).font = font
        src = subtotal_row[sec_key]
        tc = ws.cell(row=r, column=2, value=f"=B{src}")
        tc.number_format = euro_fmt
        tc.fill = fill
        tc.font = font
        cats_map = {"inkomen": INCOME_CATS, "vl": VASTE_LASTEN_CATS,
                    "dag": DAGELIJKS_CATS, "overig": OVERIG_CATS}
        pc = ws.cell(row=r, column=3, value=_pct(_total(cats_map[sec_key])))
        pc.number_format = pct_fmt
        pc.fill = fill
        pc.font = font
        row_num[0] += 1

    _hdr()

    _sec_hdr("INKOMEN", C_INC_HDR)
    for cat in INCOME_CATS:
        _data_row(cat)
    _subtotal("Totaal Inkomen", INCOME_CATS, C_INC_SUB, sec_key="inkomen")
    _blank()

    _sec_hdr("VASTE LASTEN", C_VL_HDR)
    for cat in VASTE_LASTEN_CATS:
        _data_row(cat)
    _subtotal("Totaal Vaste Lasten", VASTE_LASTEN_CATS, C_VL_SUB, sec_key="vl")
    _blank()

    _sec_hdr("DAGELIJKSE UITGAVEN", C_DAG_HDR)
    for cat in DAGELIJKS_CATS:
        _data_row(cat)
    _subtotal("Totaal Dagelijks", DAGELIJKS_CATS, C_DAG_SUB, sec_key="dag")
    _blank()

    _sec_hdr("OVERIG", C_OVR_HDR)
    for cat in OVERIG_CATS:
        _data_row(cat)
    _subtotal("Totaal Overig", OVERIG_CATS, C_OVR_SUB, sec_key="overig")
    _blank()
    _blank()

    _sec_hdr("SAMENVATTING", C_SAM_HDR)
    _sam_row("Totaal Inkomen",      "inkomen", C_INC_SUB)
    _blank()
    _sam_row("Totaal Vaste Lasten", "vl",      C_VL_SUB)
    _sam_row("Totaal Dagelijks",    "dag",     C_DAG_SUB)
    _sam_row("Totaal Overig",       "overig",  C_OVR_SUB)
    _blank()

    r     = row_num[0]
    netto = df_real["amount"].sum()
    color = C_NETTO_POS if netto >= 0 else C_NETTO_NEG
    fill  = _fill(color)
    font  = Font(bold=True, color=WHITE, name="Arial", size=11)
    ws.cell(row=r, column=1, value="NETTO").fill = fill
    ws.cell(row=r, column=1).font = font
    parts = [f"B{sam_row_map[k]}"
             for k in ("inkomen", "vl", "dag", "overig") if k in sam_row_map]
    tc = ws.cell(row=r, column=2, value=f"=SUM({','.join(parts)})")
    tc.number_format = euro_fmt
    tc.fill = fill
    tc.font = font
    pc = ws.cell(row=r, column=3, value=_pct(netto))
    pc.number_format = pct_fmt
    pc.fill = fill
    pc.font = font
    ws.row_dimensions[r].height = 18

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 16


def write_controle_sheet(ws, df, year_label=""):
    ws.title = f"Controle{' ' + year_label if year_label else ''}"
    ws.freeze_panes = "B2"
    euro_fmt = '€ #,##0;-€ #,##0'

    periods = sorted(df["month"].unique())
    df_real = df[df["category"] != "Interne Overboeking"]
    df_int  = df[df["category"] == "Interne Overboeking"]

    def _s(frame, period):
        return float(frame[frame["month"] == period]["amount"].sum())

    headers = [
        "Maand", "Inkomen (cat)", "Uitgaven (cat)", "Netto Alle", "Interne OB",
    ]

    r = 1
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=r, column=c, value=h)
        cell.fill = _fill(C_HEADER)
        cell.font = Font(bold=True, color=WHITE, name="Arial", size=9)
        cell.alignment = Alignment(horizontal="center" if c > 1 else "left")
    ws.row_dimensions[r].height = 18

    EXPENSE_CATS = VASTE_LASTEN_CATS + DAGELIJKS_CATS + OVERIG_CATS
    running = [0.0] * (len(headers) - 1)

    for period in periods:
        r += 1
        inc_cat = _s(df_real[df_real["category"].isin(INCOME_CATS)], period)
        uit_cat = _s(df_real[df_real["category"].isin(EXPENSE_CATS)], period)
        netto   = _s(df_real, period)
        interne = _s(df_int, period)

        vals = [inc_cat, uit_cat, netto, interne]
        ws.cell(row=r, column=1, value=format_period(period, "month")).font = Font(name="Arial", size=9)
        for c, v in enumerate(vals, 2):
            cell = ws.cell(row=r, column=c, value=round(v))
            cell.number_format = euro_fmt
            cell.font = Font(name="Arial", size=9)
            cell.alignment = Alignment(horizontal="right")
        for i, v in enumerate(vals):
            running[i] += v

    r += 1
    ws.cell(row=r, column=1, value="Totaal").font = Font(bold=True, name="Arial", size=9)
    for c, v in enumerate(running, 2):
        cell = ws.cell(row=r, column=c, value=round(v))
        cell.number_format = euro_fmt
        cell.font = Font(bold=True, name="Arial", size=9)
        cell.alignment = Alignment(horizontal="right")

    ws.column_dimensions["A"].width = 12
    for c in range(2, len(headers) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 15


def save_output(df, year_label=""):
    import time
    suffix = f" {year_label}" if year_label else ""
    wb = Workbook()

    t0 = time.time()
    ws_tx = wb.active
    write_transactions_sheet(ws_tx, df)
    ws_tx.title = f"Transacties{suffix}"
    print(f"  Sheet Transacties:      {time.time() - t0:.2f}s")

    t0 = time.time()
    ws_mo = wb.create_sheet()
    write_overview_sheet(ws_mo, df, group_by="month", tx_sheet_name=ws_tx.title, year_label=year_label)
    ws_mo.title = f"Maand Overzicht{suffix}"
    print(f"  Sheet Maand Overzicht:  {time.time() - t0:.2f}s")

    t0 = time.time()
    ws_js = wb.create_sheet()
    write_jaar_samenvatting_sheet(ws_js, df, year_label=year_label,
                                  tx_sheet_name=ws_tx.title)
    print(f"  Sheet Jaar Samenvatting:{time.time() - t0:.2f}s")

    t0 = time.time()
    ws_ctrl = wb.create_sheet()
    write_controle_sheet(ws_ctrl, df, year_label=year_label)
    print(f"  Sheet Controle:         {time.time() - t0:.2f}s")

    if year_label:
        filename = f"boekhouding_{year_label}.xlsx"
    else:
        now = datetime.now().strftime("%Y%m%d_%H%M")
        filename = f"boekhouding_{now}.xlsx"

    t0 = time.time()
    out_path = os.path.join(OUTPUT_DIR, filename)
    wb.save(out_path)
    print(f"  wb.save():              {time.time() - t0:.2f}s")
    print(f"Opgeslagen: {out_path}")
    return out_path
