from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from config import MAANDEN_NL

EURO   = "€ #,##0.00;-€ #,##0.00"
NUM2   = "#,##0.00;-#,##0.00"
C_DARK  = "1F4E79"
C_ALT   = "F2F2F2"
C_GREEN = "375623"
WHITE   = "FFFFFF"


def _fill(hex_c):
    return PatternFill("solid", fgColor=hex_c)


def _font(bold=False, color="000000", size=10):
    return Font(bold=bold, color=color, name="Arial", size=size)


def _hdr(ws, headers, widths):
    for c, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(1, c, h)
        cell.fill = _fill(C_DARK)
        cell.font = _font(bold=True, color=WHITE)
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"


def write_transacties(ws, df):
    headers = ["Datum", "Tijd", "Product", "ISIN", "Aantal", "Koers EUR", "Waarde EUR", "Kosten EUR", "Totaal EUR"]
    widths  = [12, 6, 42, 14, 9, 12, 12, 12, 12]
    _hdr(ws, headers, widths)
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    for i, (_, row) in enumerate(df.sort_values("date", ascending=False).iterrows(), 2):
        bg = _fill(C_ALT) if i % 2 == 0 else None
        values = [
            row["date"].strftime("%d-%m-%Y"),
            row["date"].strftime("%H:%M"),
            row["product"],
            row["isin"],
            row["aantal"],
            row["koers_eur"],
            row["waarde_eur"],
            row["kosten_eur"],
            row["totaal_eur"],
        ]
        fmts = [None, None, None, None, NUM2, EURO, EURO, EURO, EURO]
        for c, (val, fmt) in enumerate(zip(values, fmts), 1):
            cell = ws.cell(i, c, val)
            cell.font = _font()
            if bg:
                cell.fill = bg
            if fmt:
                cell.number_format = fmt
                cell.alignment = Alignment(horizontal="right")


def write_maand_overzicht(ws, df):
    headers = ["Maand", "Aankopen EUR", "Verkopen EUR", "Kosten EUR", "Netto kasstroom EUR"]
    widths  = [14, 18, 18, 14, 20]
    _hdr(ws, headers, widths)

    months = sorted(df["month"].dropna().unique())
    for i, m in enumerate(months, 2):
        mdf      = df[df["month"] == m]
        aankopen = mdf[mdf["totaal_eur"] < 0]["totaal_eur"].sum()
        verkopen = mdf[mdf["totaal_eur"] > 0]["totaal_eur"].sum()
        kosten   = mdf["kosten_eur"].sum()
        netto    = mdf["totaal_eur"].sum()
        yr, mo   = m.split("-")
        bg = _fill(C_ALT) if i % 2 == 0 else None
        for c, val in enumerate([f"{MAANDEN_NL[int(mo)]} {yr}", aankopen, verkopen, kosten, netto], 1):
            cell = ws.cell(i, c, val)
            cell.font = _font()
            if bg:
                cell.fill = bg
            if c > 1:
                cell.number_format = EURO
                cell.alignment = Alignment(horizontal="right")

    r = len(months) + 2
    totals = [
        "Totaal",
        df[df["totaal_eur"] < 0]["totaal_eur"].sum(),
        df[df["totaal_eur"] > 0]["totaal_eur"].sum(),
        df["kosten_eur"].sum(),
        df["totaal_eur"].sum(),
    ]
    for c, val in enumerate(totals, 1):
        cell = ws.cell(r, c, val)
        cell.font = _font(bold=True, color=WHITE)
        cell.fill = _fill(C_GREEN)
        if c > 1:
            cell.number_format = EURO
            cell.alignment = Alignment(horizontal="right")
    ws.row_dimensions[r].height = 20


def write_per_product(ws, df):
    headers = ["Product", "ISIN", "Aankopen EUR", "Verkopen EUR", "Kosten EUR", "Netto EUR"]
    widths  = [42, 14, 16, 16, 14, 14]
    _hdr(ws, headers, widths)

    grp = df.groupby(["product", "isin"], sort=True)
    for i, ((prod, isin), gdf) in enumerate(grp, 2):
        aankopen = gdf[gdf["totaal_eur"] < 0]["totaal_eur"].sum()
        verkopen = gdf[gdf["totaal_eur"] > 0]["totaal_eur"].sum()
        kosten   = gdf["kosten_eur"].sum()
        netto    = gdf["totaal_eur"].sum()
        bg = _fill(C_ALT) if i % 2 == 0 else None
        for c, val in enumerate([prod, isin, aankopen, verkopen, kosten, netto], 1):
            cell = ws.cell(i, c, val)
            cell.font = _font()
            if bg:
                cell.fill = bg
            if c > 2:
                cell.number_format = EURO
                cell.alignment = Alignment(horizontal="right")
