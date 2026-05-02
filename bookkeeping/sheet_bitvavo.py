from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from config import MAANDEN_NL

EURO    = "€ #,##0.00;-€ #,##0.00"
NUM8    = "#,##0.00000000"
C_DARK  = "1F4E79"
C_ALT   = "F2F2F2"
C_GREEN = "375623"
WHITE   = "FFFFFF"

TYPE_NL = {
    "buy":        "Aankoop",
    "sell":       "Verkoop",
    "deposit":    "Storting",
    "withdrawal": "Opname",
}


def _fill(hex_c):
    return PatternFill("solid", fgColor=hex_c)


def _font(bold=False, color="000000", size=10):
    return Font(bold=bold, color=color, name="Arial", size=size)


def _hdr(ws, headers, widths):
    for c, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(1, c, h)
        cell.fill = _fill(C_DARK)
        cell.font = _font(bold=True, color=WHITE)
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"


def write_transacties(ws, df):
    headers = ["Datum", "Tijd", "Type", "Munt", "Aantal", "EUR bedrag", "Kosten EUR"]
    widths  = [12, 12, 10, 8, 18, 14, 12]
    _hdr(ws, headers, widths)
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    for i, (_, row) in enumerate(df.sort_values("date", ascending=False).iterrows(), 2):
        bg = _fill(C_ALT) if i % 2 == 0 else None
        values = [
            row["date"].strftime("%d-%m-%Y"),
            row["date"].strftime("%H:%M:%S"),
            TYPE_NL.get(row["type"], row["type"]),
            row["currency"],
            row["amount"],
            row["eur_amount"],
            row["fee_eur"],
        ]
        fmts = [None, None, None, None, NUM8, EURO, EURO]
        for c, (val, fmt) in enumerate(zip(values, fmts), 1):
            cell = ws.cell(i, c, val)
            cell.font = _font()
            if bg:
                cell.fill = bg
            if fmt:
                cell.number_format = fmt
                cell.alignment = Alignment(horizontal="right")


def write_maand_overzicht(ws, df):
    headers = ["Maand", "Stortingen EUR", "Aankopen EUR", "Verkopen EUR", "Kosten EUR", "Netto EUR saldo"]
    widths  = [14, 16, 16, 16, 14, 16]
    _hdr(ws, headers, widths)

    months = sorted(df["month"].dropna().unique())
    for i, m in enumerate(months, 2):
        mdf       = df[df["month"] == m]
        stortingen = mdf[mdf["type"] == "deposit"]["eur_amount"].sum()
        aankopen   = mdf[mdf["type"] == "buy"]["eur_amount"].sum()
        verkopen   = mdf[mdf["type"] == "sell"]["eur_amount"].sum()
        kosten     = mdf["fee_eur"].sum()
        netto      = mdf["eur_amount"].sum()
        yr, mo     = m.split("-")
        bg = _fill(C_ALT) if i % 2 == 0 else None
        for c, val in enumerate([f"{MAANDEN_NL[int(mo)]} {yr}", stortingen, aankopen, verkopen, kosten, netto], 1):
            cell = ws.cell(i, c, val)
            cell.font = _font()
            if bg:
                cell.fill = bg
            if c > 1:
                cell.number_format = EURO
                cell.alignment = Alignment(horizontal="right")

    r = len(months) + 2
    totals = [
        "Totaal",
        df[df["type"] == "deposit"]["eur_amount"].sum(),
        df[df["type"] == "buy"]["eur_amount"].sum(),
        df[df["type"] == "sell"]["eur_amount"].sum(),
        df["fee_eur"].sum(),
        df["eur_amount"].sum(),
    ]
    for c, val in enumerate(totals, 1):
        cell = ws.cell(r, c, val)
        cell.font = _font(bold=True, color=WHITE)
        cell.fill = _fill(C_GREEN)
        if c > 1:
            cell.number_format = EURO
            cell.alignment = Alignment(horizontal="right")
    ws.row_dimensions[r].height = 20


def write_per_coin(ws, df):
    headers = ["Munt", "Stortingen EUR", "Aankopen EUR", "Verkopen EUR", "Kosten EUR", "Netto EUR"]
    widths  = [10, 16, 16, 16, 14, 14]
    _hdr(ws, headers, widths)

    for i, (coin, gdf) in enumerate(sorted(df.groupby("currency")), 2):
        stortingen = gdf[gdf["type"] == "deposit"]["eur_amount"].sum()
        aankopen   = gdf[gdf["type"] == "buy"]["eur_amount"].sum()
        verkopen   = gdf[gdf["type"] == "sell"]["eur_amount"].sum()
        kosten     = gdf["fee_eur"].sum()
        netto      = gdf["eur_amount"].sum()
        bg = _fill(C_ALT) if i % 2 == 0 else None
        for c, val in enumerate([coin, stortingen, aankopen, verkopen, kosten, netto], 1):
            cell = ws.cell(i, c, val)
            cell.font = _font()
            if bg:
                cell.fill = bg
            if c > 1:
                cell.number_format = EURO
                cell.alignment = Alignment(horizontal="right")
