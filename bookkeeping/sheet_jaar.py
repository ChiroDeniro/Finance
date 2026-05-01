from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

from config import (
    INCOME_CATS, VASTE_LASTEN_CATS, DAGELIJKS_CATS, OVERIG_CATS,
    WHITE,
    C_HEADER,
    C_INC_HDR, C_INC_SUB,
    C_VL_HDR,  C_VL_SUB,
    C_DAG_HDR, C_DAG_SUB,
    C_OVR_HDR, C_OVR_SUB,
    C_SAM_HDR, C_NETTO_POS, C_NETTO_NEG,
    _fill,
)


def write_jaar_samenvatting_sheet(ws, df, year_label="", tx_sheet_name="Transacties"):
    ws.title = f"Jaar Samenvatting{' ' + year_label if year_label else ''}"
    ws.freeze_panes = "B2"

    euro_fmt = '€ #,##0;-€ #,##0'
    pct_fmt  = '0.0"%"'

    EXCLUDE = {"Interne Overboeking"}
    df_real = df[~df["category"].isin(EXCLUDE)]
    cat_totals = df_real.groupby("category")["amount"].sum()

    def _total(cats):
        return sum(cat_totals.get(c, 0.0) for c in cats)

    totaal_inkomen = _total(INCOME_CATS)

    def _pct(val):
        return round(val / totaal_inkomen * 100, 1) if totaal_inkomen else 0.0

    tx = f"'{tx_sheet_name}'" if ' ' in tx_sheet_name else tx_sheet_name

    def _cat_formula(cat):
        if year_label:
            yr = int(year_label)
            return (f'=SUMIFS({tx}!$E:$E,{tx}!$F:$F,"{cat}",'
                    f'{tx}!$A:$A,">="&DATE({yr},1,1),'
                    f'{tx}!$A:$A,"<"&DATE({yr + 1},1,1))')
        return f'=SUMIF({tx}!$F:$F,"{cat}",{tx}!$E:$E)'

    row_num      = [2]
    row_for_cat  = {}
    subtotal_row = {}
    sam_row_map  = {}

    def _hdr():
        r = row_num[0]
        for c, h in enumerate(["Categorie", "Totaal", "% van Inkomen"], 1):
            cell = ws.cell(row=r, column=c, value=h)
            cell.fill = _fill(C_HEADER)
            cell.font = Font(bold=True, color=WHITE, name="Arial", size=10)
            cell.alignment = Alignment(horizontal="center" if c > 1 else "left")
        ws.row_dimensions[r].height = 20
        row_num[0] += 1

    def _sec_hdr(label, color):
        r = row_num[0]
        for c in range(1, 4):
            ws.cell(row=r, column=c).fill = _fill(color)
            ws.cell(row=r, column=c).font = Font(bold=True, color=WHITE, name="Arial", size=9)
        ws.cell(row=r, column=1).value = label
        ws.row_dimensions[r].height = 16
        row_num[0] += 1

    def _data_row(cat, _unused=None):
        r = row_num[0]
        row_for_cat[cat] = r
        ws.cell(row=r, column=1, value=cat)
        ws.cell(row=r, column=1).font = Font(name="Arial", size=9)
        tc = ws.cell(row=r, column=2, value=_cat_formula(cat))
        tc.number_format = euro_fmt
        tc.font = Font(name="Arial", size=9)
        tc.alignment = Alignment(horizontal="right")
        pc = ws.cell(row=r, column=3, value=_pct(cat_totals.get(cat, 0.0)))
        pc.number_format = pct_fmt
        pc.font = Font(italic=True, name="Arial", size=9)
        pc.alignment = Alignment(horizontal="right")
        row_num[0] += 1

    def _subtotal(label, cats, color, sec_key=None):
        r = row_num[0]
        if sec_key:
            subtotal_row[sec_key] = r
        fill = _fill(color)
        font = Font(bold=True, color=WHITE, name="Arial", size=9)
        ws.cell(row=r, column=1, value=label).fill = fill
        ws.cell(row=r, column=1).font = font
        parts = [f"B{row_for_cat[c]}" for c in cats if c in row_for_cat]
        tc = ws.cell(row=r, column=2,
                     value=f"=SUM({','.join(parts)})" if parts else "=0")
        tc.number_format = euro_fmt
        tc.fill = fill
        tc.font = font
        pc = ws.cell(row=r, column=3, value=_pct(_total(cats)))
        pc.number_format = pct_fmt
        pc.fill = fill
        pc.font = font
        ws.row_dimensions[r].height = 15
        row_num[0] += 1

    def _blank():
        ws.row_dimensions[row_num[0]].height = 5
        row_num[0] += 1

    def _sam_row(label, sec_key, sub_color):
        r = row_num[0]
        sam_row_map[sec_key] = r
        fill = _fill(sub_color)
        font = Font(bold=True, color=WHITE, name="Arial", size=9)
        ws.cell(row=r, column=1, value=label).fill = fill
        ws.cell(row=r, column=1).font = font
        src = subtotal_row[sec_key]
        tc = ws.cell(row=r, column=2, value=f"=B{src}")
        tc.number_format = euro_fmt
        tc.fill = fill
        tc.font = font
        cats_map = {"inkomen": INCOME_CATS, "vl": VASTE_LASTEN_CATS,
                    "dag": DAGELIJKS_CATS, "overig": OVERIG_CATS}
        pc = ws.cell(row=r, column=3, value=_pct(_total(cats_map[sec_key])))
        pc.number_format = pct_fmt
        pc.fill = fill
        pc.font = font
        row_num[0] += 1

    _hdr()

    _sec_hdr("INKOMEN", C_INC_HDR)
    for cat in INCOME_CATS:
        _data_row(cat)
    _subtotal("Totaal Inkomen", INCOME_CATS, C_INC_SUB, sec_key="inkomen")
    _blank()

    _sec_hdr("VASTE LASTEN", C_VL_HDR)
    for cat in VASTE_LASTEN_CATS:
        _data_row(cat)
    _subtotal("Totaal Vaste Lasten", VASTE_LASTEN_CATS, C_VL_SUB, sec_key="vl")
    _blank()

    _sec_hdr("DAGELIJKSE UITGAVEN", C_DAG_HDR)
    for cat in DAGELIJKS_CATS:
        _data_row(cat)
    _subtotal("Totaal Dagelijks", DAGELIJKS_CATS, C_DAG_SUB, sec_key="dag")
    _blank()

    _sec_hdr("OVERIG", C_OVR_HDR)
    for cat in OVERIG_CATS:
        _data_row(cat)
    _subtotal("Totaal Overig", OVERIG_CATS, C_OVR_SUB, sec_key="overig")
    _blank()
    _blank()

    _sec_hdr("SAMENVATTING", C_SAM_HDR)
    _sam_row("Totaal Inkomen",      "inkomen", C_INC_SUB)
    _blank()
    _sam_row("Totaal Vaste Lasten", "vl",      C_VL_SUB)
    _sam_row("Totaal Dagelijks",    "dag",     C_DAG_SUB)
    _sam_row("Totaal Overig",       "overig",  C_OVR_SUB)
    _blank()

    r     = row_num[0]
    netto = df_real["amount"].sum()
    color = C_NETTO_POS if netto >= 0 else C_NETTO_NEG
    fill  = _fill(color)
    font  = Font(bold=True, color=WHITE, name="Arial", size=11)
    ws.cell(row=r, column=1, value="NETTO").fill = fill
    ws.cell(row=r, column=1).font = font
    parts = [f"B{sam_row_map[k]}"
             for k in ("inkomen", "vl", "dag", "overig") if k in sam_row_map]
    tc = ws.cell(row=r, column=2, value=f"=SUM({','.join(parts)})")
    tc.number_format = euro_fmt
    tc.fill = fill
    tc.font = font
    pc = ws.cell(row=r, column=3, value=_pct(netto))
    pc.number_format = pct_fmt
    pc.fill = fill
    pc.font = font
    ws.row_dimensions[r].height = 18

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 16
