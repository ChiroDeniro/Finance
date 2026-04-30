"""
ABN AMRO Bookkeeping Processor
================================
Drop your ABN AMRO Excel (TXT) export into the /input folder, then run:
    python process.py

Output: /output/boekhouding_YYYYMMDD_HHMM.xlsx  with:
  - Sheet 1: Transacties       — all transactions, sorted by date
  - Sheet 2: Maand Overzicht   — kasboek-style blocks per month
  - Sheet 3: Jaar Overzicht    — same structure per year
  - Sheet 4: Onbekend          — unknown merchants grouped for easy triage
"""

import os
import re
import sys
import glob
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# ── Paths ─────────────────────────────────────────────────────────────────────
BASE_DIR   = os.path.dirname(os.path.abspath(__file__))
INPUT_DIR  = os.path.join(BASE_DIR, "input")
RULES_FILE = os.path.join(BASE_DIR, "rules.xlsx")

# Save to Drive-synced folder if it exists, otherwise local output/
_DRIVE_DIR = r"C:\Users\chris\Documents\Finance\Kasboek"
OUTPUT_DIR = _DRIVE_DIR if os.path.isdir(_DRIVE_DIR) else os.path.join(BASE_DIR, "output")

os.makedirs(INPUT_DIR,  exist_ok=True)
os.makedirs(OUTPUT_DIR, exist_ok=True)

# ── Category definitions ──────────────────────────────────────────────────────
INCOME_CATS = [
    "Salaris",
    "DUO / Studiefinanciering",
    "Zorgtoeslag",
    "Familie & Giften",
    "Inkomsten Overig",
]
VASTE_LASTEN_CATS = [
    "Huur",
    "Inclusief Huur",
    "Zorgverzekering",
    "Telefoon & Internet",
    "Bankkosten",
    "Abonnementen",
    "Sport & Fitness",
    "Onderhoud",
]
DAGELIJKS_CATS = [
    "Boodschappen",
    "Eten & Drinken",
    "Uitgaan",
    "OV & Reizen",
    "Kleding",
    "Kapper",
    "Gezondheid",
    "WbW",
    "Cultuur & Entertainment",
    "Studie",
    "Dagelijks Overig",
]
OVERIG_CATS = ["Sparen", "Diversen"]

ALL_KNOWN_CATS = INCOME_CATS + VASTE_LASTEN_CATS + DAGELIJKS_CATS + OVERIG_CATS

OWN_ACCOUNTS = {"536542171", "844835730"}

# ── Colors ────────────────────────────────────────────────────────────────────
WHITE = "FFFFFF"

C_INC_HDR  = "375623"   # dark forest green
C_INC_SUB  = "548235"   # medium green
C_INC_ROW  = "E2EFDA"   # light green

C_VL_HDR   = "1F4E79"   # dark blue
C_VL_SUB   = "2E75B6"   # medium blue
C_VL_ROW   = "DEEAF1"   # light blue

C_DAG_HDR  = "843C0C"   # dark orange-brown
C_DAG_SUB  = "C55A11"   # medium orange
C_DAG_ROW  = "FCE4D6"   # light orange

C_OVR_HDR  = "595959"   # dark gray
C_OVR_SUB  = "808080"   # medium gray
C_OVR_ROW  = "F2F2F2"   # light gray

C_SAM_HDR  = "1F4E79"   # dark blue (same as VL)
C_NETTO_POS = "375623"  # green
C_NETTO_NEG = "C00000"  # red
C_KOSTEN_ROW = "FFD6E0" # lichtroze voor Totaal Kosten

C_HEADER   = "1F4E79"
C_ALT_ROW  = "EEF3FA"
C_UNKNOWN  = "FFF2CC"

# ── Dutch month labels ────────────────────────────────────────────────────────
MAANDEN_NL = {
    1: "Jan", 2: "Feb", 3: "Mrt", 4: "Apr", 5: "Mei",  6: "Jun",
    7: "Jul", 8: "Aug", 9: "Sep", 10: "Okt", 11: "Nov", 12: "Dec",
}


def format_period(period_str, group_by):
    """'2026-03' → "Mrt '26"  (month)  or  '2026'  (year)"""
    if group_by == "year":
        return period_str
    yr, mo = period_str.split("-")
    return f"{MAANDEN_NL[int(mo)]} '{yr[2:]}"


def dutch_euros(value):
    """€2,035 → '€2.035'  (Dutch thousands separator for terminal output)"""
    return "€" + f"{int(round(value)):,}".replace(",", ".")


# ── Step 1: Find & load TAB files ─────────────────────────────────────────────
def find_input_files():
    patterns = ["*.TAB", "*.tab", "*.txt", "*.TXT"]
    seen, files = set(), []
    for p in patterns:
        for f in glob.glob(os.path.join(INPUT_DIR, p)):
            key = os.path.normcase(os.path.abspath(f))
            if key not in seen:
                seen.add(key)
                files.append(f)
    if not files:
        print(f"Geen invoerbestanden gevonden in {INPUT_DIR}/")
        print("Download je ABN AMRO afschrift als 'Excel (TXT)' en zet het daar neer.")
        sys.exit(1)
    print(f"Gevonden: {len(files)} invoerbestand(en)")
    for f in files:
        print(f"  {os.path.basename(f)}")
    return files


def validate_tab_file(df, filepath):
    """Validate ABN AMRO TAB export format; exit with clear Dutch error if wrong."""
    name   = os.path.basename(filepath)
    errors = []

    if len(df.columns) != 8:
        errors.append(f"verwacht 8 kolommen, gevonden {len(df.columns)}")

    if len(df) == 0:
        errors.append("bestand bevat geen transacties")
    else:
        for v in df["date"].dropna().head(3):
            if not re.match(r"^\d{8}$", str(v).strip()):
                errors.append(
                    f"datum-kolom niet in JJJJMMDD-formaat (voorbeeld: '{v}')"
                )
                break
        if not any("," in str(v) for v in df["amount"].dropna().head(5)):
            errors.append(
                "bedrag-kolom mist komma als decimaalteken "
                "(verwacht bijv. '-19,42')"
            )

    if errors:
        print(f"\nOngeldig bestandsformaat: {name}")
        for e in errors:
            print(f"  - {e}")
        print("Verwacht: ABN AMRO 'Excel (TXT)' export, 8 tab-kolommen:")
        print("  rekening | valuta | datum (JJJJMMDD) | saldo_voor | saldo_na"
              " | valutadatum | bedrag | omschrijving")
        sys.exit(1)


def load_transactions(files):
    dfs = []
    for f in files:
        raw = pd.read_csv(
            f, sep="\t", header=None, encoding="utf-8",
            names=["account", "currency", "date", "balance_before",
                   "balance_after", "value_date", "amount", "description"],
            dtype=str,
        )
        validate_tab_file(raw, f)
        dfs.append(raw)

    df = pd.concat(dfs, ignore_index=True).drop_duplicates()
    df["date"]        = pd.to_datetime(df["date"], format="%Y%m%d")
    df["month"]       = df["date"].dt.to_period("M").astype(str)
    df["year"]        = df["date"].dt.year.astype(str)
    df["amount"]      = df["amount"].str.replace(",", ".").astype(float)
    df["description"] = df["description"].str.strip()
    df["merchant"]    = df["description"].apply(extract_merchant)
    df = df.sort_values("date").reset_index(drop=True)

    print(f"Geladen: {len(df)} transacties  "
          f"({df['date'].min().date()} - {df['date'].max().date()})")
    return df


def extract_merchant(desc):
    desc = str(desc).strip()
    # BEA / Apple Pay pin transaction
    m = re.search(r"BEA,.*?\s{2,}(.+?),PAS", desc)
    if m:
        return m.group(1).strip()
    # Standard SEPA /NAME/ format
    m = re.search(r"/NAME/([^/]+)", desc)
    if m:
        return m.group(1).strip()
    # SEPA Incasso "Naam: ..." format (e.g. NS GROEP, gym subscriptions)
    m = re.search(r"Naam:\s*([^\t\n/,]+)", desc)
    if m:
        return m.group(1).strip()
    if "ABN AMRO" in desc.upper():
        return "ABN AMRO Bank"
    return desc[:40]


# ── Step 2: Load categorisation rules ────────────────────────────────────────
def load_rules():
    if not os.path.exists(RULES_FILE):
        print(f"Geen rules.xlsx gevonden op {RULES_FILE}")
        print("Voer uit met --create-rules om een startbestand te maken.")
        return []

    df = pd.read_excel(RULES_FILE, sheet_name="Rules")
    rules = []
    for _, row in df.iterrows():
        kw  = str(row.get("keyword",  "")).strip()
        cat = str(row.get("category", "")).strip()
        if kw and cat:
            rules.append((kw.lower(), cat))
    print(f"Geladen: {len(rules)} categorisatieregels")
    return rules


def categorise(merchant, rules):
    ml = merchant.lower()
    for kw, cat in rules:
        if kw in ml:
            return cat
    return "Onbekend"


def apply_categories(df, rules):
    df["category"] = df["merchant"].apply(lambda m: categorise(m, rules))
    n_unk   = (df["category"] == "Onbekend").sum()
    n_total = len(df)
    pct     = 100 * (n_total - n_unk) / n_total if n_total else 0
    print(f"Gecategoriseerd: {n_total - n_unk}/{n_total}  ({pct:.0f}%)")
    if n_unk:
        print(f"  {n_unk} transacties zonder categorie - zie het 'Onbekend' tabblad")
    return df


def detect_internal_transfers(df):
    """Mark transfers between own accounts as 'Interne Overboeking'."""
    mask = pd.Series(False, index=df.index)
    for (date_val, abs_amt), group in df.groupby(
        [df["date"].dt.date, df["amount"].abs()]
    ):
        if len(group) >= 2:
            accs = set(group["account"].unique())
            if len(accs.intersection(OWN_ACCOUNTS)) >= 2:
                if abs(group["amount"].sum()) < 0.02:
                    mask.loc[group.index] = True
    n = int(mask.sum())
    if n:
        df.loc[mask, "category"] = "Interne Overboeking"
        print(f"Interne overboekingen: {n} transacties "
              f"(betaalrekening <-> spaarrekening, uitgesloten van netto)")
    return df


# ── Step 3: Build output Excel ────────────────────────────────────────────────

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def write_transactions_sheet(ws, df):
    ws.title = "Transacties"
    ws.freeze_panes = "A2"

    headers = ["Datum", "Rekening", "Omschrijving", "Merchant",
               "Bedrag (EUR)", "Categorie", "Maand"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = _fill(C_HEADER)
        cell.font = Font(bold=True, color=WHITE, name="Arial", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 20

    unk_fill = _fill(C_UNKNOWN)
    inc_fill = _fill(C_INC_ROW)
    alt_fill = _fill(C_ALT_ROW)
    euro_fmt = '€ #,##0.00;-€ #,##0.00'

    for r, (_, row) in enumerate(df.iterrows(), 2):
        is_unk    = row["category"] == "Onbekend"
        is_income = row["amount"] > 0
        fill = (unk_fill if is_unk
                else (inc_fill if is_income
                      else (alt_fill if r % 2 == 0 else None)))
        data = [
            row["date"].date(), row["account"], row["description"][:80],
            row["merchant"], row["amount"], row["category"], row["month"],
        ]
        for c, val in enumerate(data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = Font(name="Arial", size=9)
            cell.alignment = Alignment(vertical="center")
            if fill:
                cell.fill = fill
            if c == 5:
                cell.number_format = euro_fmt

    for c, w in enumerate([12, 12, 50, 28, 14, 22, 10], 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


def write_overview_sheet(ws, df, group_by="month"):
    """
    Writes a kasboek-style overview with four colour-coded blocks
    (INKOMEN / VASTE LASTEN / DAGELIJKSE UITGAVEN / OVERIG) followed by
    a SAMENVATTING block and a green/red NETTO row.

    group_by="month"  →  columns are Dutch short month names + Totaal + Gemiddeld
    group_by="year"   →  columns are years + Totaal
    """
    if group_by == "month":
        ws.title    = "Maand Overzicht"
        periods     = sorted(df["month"].unique())
        pivot_col   = "month"
    else:
        ws.title    = "Jaar Overzicht"
        periods     = sorted(df["year"].unique())
        pivot_col   = "year"

    ws.freeze_panes = "B2"

    n          = len(periods)
    tot_col    = n + 2
    gem_col    = n + 3 if group_by == "month" else None
    last_col   = gem_col or tot_col
    euro_fmt   = '€ #,##0;-€ #,##0'

    # Pivot for lookups (exclude non-real categories)
    EXCLUDE_FROM_OVERVIEW = {"Onbekend", "Interne Overboeking"}
    pivot = df[~df["category"].isin(EXCLUDE_FROM_OVERVIEW)].pivot_table(
        index="category", columns=pivot_col,
        values="amount", aggfunc="sum", fill_value=0,
    )

    def _val(cat, period):
        try:
            return float(pivot.loc[cat, period])
        except KeyError:
            return 0.0

    # ── Row pointer (shared across all nested helpers) ────────────────────────
    row_num = [2]   # list so nested functions can mutate it

    # ── Helpers ───────────────────────────────────────────────────────────────
    def _header_row():
        r = row_num[0]
        ws.cell(row=r, column=1, value="Categorie")
        ws.cell(row=r, column=1).fill = _fill(C_HEADER)
        ws.cell(row=r, column=1).font = Font(bold=True, color=WHITE, name="Arial", size=10)
        for c, p in enumerate(periods, 2):
            cell = ws.cell(row=r, column=c, value=format_period(p, group_by))
            cell.fill = _fill(C_HEADER)
            cell.font = Font(bold=True, color=WHITE, name="Arial", size=10)
            cell.alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=tot_col, value="Totaal")
        ws.cell(row=r, column=tot_col).fill = _fill(C_HEADER)
        ws.cell(row=r, column=tot_col).font = Font(bold=True, color=WHITE, name="Arial", size=10)
        ws.cell(row=r, column=tot_col).alignment = Alignment(horizontal="center")
        if gem_col:
            ws.cell(row=r, column=gem_col, value="Gemiddeld")
            ws.cell(row=r, column=gem_col).fill = _fill(C_HEADER)
            ws.cell(row=r, column=gem_col).font = Font(bold=True, color=WHITE, name="Arial", size=10)
            ws.cell(row=r, column=gem_col).alignment = Alignment(horizontal="center")
        ws.row_dimensions[r].height = 20
        row_num[0] += 1

    def _section_hdr(label, color):
        r = row_num[0]
        for c in range(1, last_col + 1):
            ws.cell(row=r, column=c).fill = _fill(color)
            ws.cell(row=r, column=c).font = Font(bold=True, color=WHITE, name="Arial", size=9)
        ws.cell(row=r, column=1).value = label
        ws.row_dimensions[r].height = 16
        row_num[0] += 1

    def _data_row(cat, _unused_color=None):
        r = row_num[0]
        ws.cell(row=r, column=1, value=cat)
        ws.cell(row=r, column=1).font = Font(name="Arial", size=9)
        total = 0.0
        for c, p in enumerate(periods, 2):
            v    = _val(cat, p)
            cell = ws.cell(row=r, column=c, value=round(v))
            cell.number_format = euro_fmt
            cell.font = Font(name="Arial", size=9)
            cell.alignment = Alignment(horizontal="right")
            total += v
        tc = ws.cell(row=r, column=tot_col, value=round(total))
        tc.number_format = euro_fmt
        tc.font = Font(bold=True, name="Arial", size=9)
        tc.alignment = Alignment(horizontal="right")
        if gem_col:
            avg = total / n if n else 0
            gc = ws.cell(row=r, column=gem_col, value=round(avg))
            gc.number_format = euro_fmt
            gc.font = Font(italic=True, name="Arial", size=9)
            gc.alignment = Alignment(horizontal="right")
        row_num[0] += 1

    def _subtotal_row(label, cats, color):
        r     = row_num[0]
        fill  = _fill(color)
        font  = Font(bold=True, color=WHITE, name="Arial", size=9)
        ws.cell(row=r, column=1, value=label).fill = fill
        ws.cell(row=r, column=1).font = font
        grand = 0.0
        for c, p in enumerate(periods, 2):
            v    = sum(_val(cat, p) for cat in cats)
            cell = ws.cell(row=r, column=c, value=round(v))
            cell.number_format = euro_fmt
            cell.fill = fill
            cell.font = font
            grand += v
        tc = ws.cell(row=r, column=tot_col, value=round(grand))
        tc.number_format = euro_fmt
        tc.fill = fill
        tc.font = font
        if gem_col:
            avg = grand / n if n else 0
            gc = ws.cell(row=r, column=gem_col, value=round(avg))
            gc.number_format = euro_fmt
            gc.fill = fill
            gc.font = font
        ws.row_dimensions[r].height = 15
        row_num[0] += 1

    def _blank():
        ws.row_dimensions[row_num[0]].height = 5
        row_num[0] += 1

    def _unk_row(label, series):
        r = row_num[0]
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=1).font = Font(name="Arial", size=9, italic=True, color="888888")
        total = 0.0
        for c, p in enumerate(periods, 2):
            v    = float(series.get(p, 0.0))
            cell = ws.cell(row=r, column=c, value=round(v) if v else 0)
            cell.number_format = euro_fmt
            cell.font = Font(name="Arial", size=9, italic=True, color="888888")
            cell.alignment = Alignment(horizontal="right")
            total += v
        tc = ws.cell(row=r, column=tot_col, value=round(total))
        tc.number_format = euro_fmt
        tc.font = Font(bold=True, name="Arial", size=9, italic=True, color="888888")
        tc.alignment = Alignment(horizontal="right")
        if gem_col:
            avg = total / n if n else 0
            gc = ws.cell(row=r, column=gem_col, value=round(avg))
            gc.number_format = euro_fmt
            gc.font = Font(italic=True, name="Arial", size=9, color="888888")
            gc.alignment = Alignment(horizontal="right")
        row_num[0] += 1

    def _samenvatting_row(label, cats, row_color, sub_color):
        """A summary row with subtotal-style bold text and colored fill."""
        r    = row_num[0]
        fill = _fill(sub_color)
        font = Font(bold=True, color=WHITE, name="Arial", size=9)
        ws.cell(row=r, column=1, value=label).fill = fill
        ws.cell(row=r, column=1).font = font
        grand = 0.0
        for c, p in enumerate(periods, 2):
            v    = sum(_val(cat, p) for cat in cats)
            cell = ws.cell(row=r, column=c, value=round(v))
            cell.number_format = euro_fmt
            cell.fill = fill
            cell.font = font
            grand += v
        tc = ws.cell(row=r, column=tot_col, value=round(grand))
        tc.number_format = euro_fmt
        tc.fill = fill
        tc.font = font
        if gem_col:
            avg = grand / n if n else 0
            gc = ws.cell(row=r, column=gem_col, value=round(avg))
            gc.number_format = euro_fmt
            gc.fill = fill
            gc.font = font
        row_num[0] += 1

    # ═══════════════════════════════════════════════════════════════════════════
    # Column headers
    # ═══════════════════════════════════════════════════════════════════════════
    _header_row()

    # ═══════════════════════════════════════════════════════════════════════════
    # INKOMEN
    # ═══════════════════════════════════════════════════════════════════════════
    _section_hdr("INKOMEN", C_INC_HDR)
    for cat in INCOME_CATS:
        _data_row(cat, C_INC_ROW)
    _subtotal_row("Totaal Inkomen", INCOME_CATS, C_INC_SUB)
    _blank()

    # ═══════════════════════════════════════════════════════════════════════════
    # VASTE LASTEN
    # ═══════════════════════════════════════════════════════════════════════════
    _section_hdr("VASTE LASTEN", C_VL_HDR)
    for cat in VASTE_LASTEN_CATS:
        _data_row(cat, C_VL_ROW)
    _subtotal_row("Totaal Vaste Lasten", VASTE_LASTEN_CATS, C_VL_SUB)
    _blank()

    # ═══════════════════════════════════════════════════════════════════════════
    # DAGELIJKSE UITGAVEN
    # ═══════════════════════════════════════════════════════════════════════════
    _section_hdr("DAGELIJKSE UITGAVEN", C_DAG_HDR)
    for cat in DAGELIJKS_CATS:
        _data_row(cat, C_DAG_ROW)
    _subtotal_row("Totaal Dagelijks", DAGELIJKS_CATS, C_DAG_SUB)
    _blank()

    # ═══════════════════════════════════════════════════════════════════════════
    # OVERIG
    # ═══════════════════════════════════════════════════════════════════════════
    _section_hdr("OVERIG", C_OVR_HDR)
    for cat in OVERIG_CATS:
        _data_row(cat, C_OVR_ROW)
    _subtotal_row("Totaal Overig", OVERIG_CATS, C_OVR_SUB)
    _blank()

    # ═══════════════════════════════════════════════════════════════════════════
    # ONBEKEND
    # ═══════════════════════════════════════════════════════════════════════════
    unk_df  = df[df["category"] == "Onbekend"]
    unk_inc = unk_df[unk_df["amount"] > 0].groupby(pivot_col)["amount"].sum()
    unk_uit = unk_df[unk_df["amount"] < 0].groupby(pivot_col)["amount"].sum()
    if len(unk_df) > 0:
        _section_hdr("ONBEKEND", C_OVR_HDR)
        _unk_row("Onbekend Inkomen", unk_inc)
        _unk_row("Onbekend Uitgaven", unk_uit)
        _blank()
    _blank()

    # ═══════════════════════════════════════════════════════════════════════════
    # SAMENVATTING
    # ═══════════════════════════════════════════════════════════════════════════
    _section_hdr("SAMENVATTING", C_SAM_HDR)
    _samenvatting_row("Totaal Inkomen",      INCOME_CATS,       C_INC_ROW, C_INC_SUB)
    _samenvatting_row("Totaal Vaste Lasten", VASTE_LASTEN_CATS, C_VL_ROW,  C_VL_SUB)
    _samenvatting_row("Totaal Dagelijks",    DAGELIJKS_CATS,    C_DAG_ROW, C_DAG_SUB)
    _samenvatting_row("Totaal Overig",       OVERIG_CATS,       C_OVR_ROW, C_OVR_SUB)

    # TOTAAL KOSTEN — alle uitgaven opgeteld als positief bedrag
    EXPENSE_CATS = VASTE_LASTEN_CATS + DAGELIJKS_CATS + OVERIG_CATS
    r    = row_num[0]
    fill = _fill(C_KOSTEN_ROW)
    font = Font(bold=True, name="Arial", size=9)
    ws.cell(row=r, column=1, value="Totaal Kosten").fill = fill
    ws.cell(row=r, column=1).font = font
    grand = 0.0
    for c, p in enumerate(periods, 2):
        v    = abs(sum(_val(cat, p) for cat in EXPENSE_CATS))
        cell = ws.cell(row=r, column=c, value=round(v))
        cell.number_format = euro_fmt
        cell.fill = fill
        cell.font = font
        grand += v
    tc = ws.cell(row=r, column=tot_col, value=round(grand))
    tc.number_format = euro_fmt
    tc.fill = fill
    tc.font = font
    if gem_col:
        gc = ws.cell(row=r, column=gem_col, value=round(grand / n if n else 0))
        gc.number_format = euro_fmt
        gc.fill = fill
        gc.font = font
    ws.row_dimensions[r].height = 15
    row_num[0] += 1

    # NETTO — green if >= 0, red if < 0, per period (exclude internal transfers)
    df_real = df[df["category"] != "Interne Overboeking"]
    r    = row_num[0]
    font = Font(bold=True, color=WHITE, name="Arial", size=11)
    ws.cell(row=r, column=1, value="NETTO").font = font
    ws.row_dimensions[r].height = 18
    netto_grand = 0.0
    for c, p in enumerate(periods, 2):
        v     = df_real[df_real[pivot_col] == p]["amount"].sum()
        color = C_NETTO_POS if v >= 0 else C_NETTO_NEG
        cell  = ws.cell(row=r, column=c, value=round(v))
        cell.number_format = euro_fmt
        cell.fill = _fill(color)
        cell.font = font
        netto_grand += v
    netto_color = C_NETTO_POS if netto_grand >= 0 else C_NETTO_NEG
    ws.cell(row=r, column=1).fill = _fill(netto_color)
    tc = ws.cell(row=r, column=tot_col, value=round(netto_grand))
    tc.number_format = euro_fmt
    tc.fill = _fill(netto_color)
    tc.font = font
    if gem_col:
        avg   = netto_grand / n if n else 0
        color = C_NETTO_POS if avg >= 0 else C_NETTO_NEG
        gc    = ws.cell(row=r, column=gem_col, value=round(avg))
        gc.number_format = euro_fmt
        gc.fill = _fill(color)
        gc.font = font

    # ── Column widths ─────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 26
    for c in range(2, last_col + 1):
        ws.column_dimensions[get_column_letter(c)].width = 12


def write_jaar_samenvatting_sheet(ws, df, year_label=""):
    """
    Single-year summary: Totaal (EUR) + % van Inkomen per category.
    Same 4-block colour structure as Maand Overzicht, but only 2 data columns.
    """
    ws.title = f"Jaar Samenvatting{' ' + year_label if year_label else ''}"
    ws.freeze_panes = "B2"

    euro_fmt = '€ #,##0;-€ #,##0'
    pct_fmt  = '0.0"%"'

    EXCLUDE = {"Onbekend", "Interne Overboeking"}
    df_real = df[~df["category"].isin(EXCLUDE)]

    cat_totals = df_real.groupby("category")["amount"].sum()

    def _total(cats):
        return sum(cat_totals.get(c, 0.0) for c in cats)

    totaal_inkomen = _total(INCOME_CATS)

    def _pct(val):
        return round(val / totaal_inkomen * 100, 1) if totaal_inkomen else 0.0

    row_num = [2]

    def _hdr():
        r = row_num[0]
        for c, h in enumerate(["Categorie", "Totaal", "% van Inkomen"], 1):
            cell = ws.cell(row=r, column=c, value=h)
            cell.fill = _fill(C_HEADER)
            cell.font = Font(bold=True, color=WHITE, name="Arial", size=10)
            cell.alignment = Alignment(horizontal="center" if c > 1 else "left")
        ws.row_dimensions[r].height = 20
        row_num[0] += 1

    def _sec_hdr(label, color):
        r = row_num[0]
        for c in range(1, 4):
            ws.cell(row=r, column=c).fill = _fill(color)
            ws.cell(row=r, column=c).font = Font(bold=True, color=WHITE,
                                                  name="Arial", size=9)
        ws.cell(row=r, column=1).value = label
        ws.row_dimensions[r].height = 16
        row_num[0] += 1

    def _data_row(cat, _unused_color=None):
        r   = row_num[0]
        val = cat_totals.get(cat, 0.0)
        ws.cell(row=r, column=1, value=cat)
        ws.cell(row=r, column=1).font = Font(name="Arial", size=9)
        tc = ws.cell(row=r, column=2, value=round(val))
        tc.number_format = euro_fmt
        tc.font = Font(name="Arial", size=9)
        tc.alignment = Alignment(horizontal="right")
        pc = ws.cell(row=r, column=3, value=_pct(val))
        pc.number_format = pct_fmt
        pc.font = Font(italic=True, name="Arial", size=9)
        pc.alignment = Alignment(horizontal="right")
        row_num[0] += 1

    def _subtotal(label, cats, color):
        r    = row_num[0]
        fill = _fill(color)
        font = Font(bold=True, color=WHITE, name="Arial", size=9)
        val  = _total(cats)
        ws.cell(row=r, column=1, value=label).fill = fill
        ws.cell(row=r, column=1).font = font
        tc = ws.cell(row=r, column=2, value=round(val))
        tc.number_format = euro_fmt
        tc.fill = fill
        tc.font = font
        pc = ws.cell(row=r, column=3, value=_pct(val))
        pc.number_format = pct_fmt
        pc.fill = fill
        pc.font = font
        ws.row_dimensions[r].height = 15
        row_num[0] += 1

    def _blank():
        ws.row_dimensions[row_num[0]].height = 5
        row_num[0] += 1

    def _sam_row(label, cats, sub_color):
        r    = row_num[0]
        fill = _fill(sub_color)
        font = Font(bold=True, color=WHITE, name="Arial", size=9)
        val  = _total(cats)
        ws.cell(row=r, column=1, value=label).fill = fill
        ws.cell(row=r, column=1).font = font
        tc = ws.cell(row=r, column=2, value=round(val))
        tc.number_format = euro_fmt
        tc.fill = fill
        tc.font = font
        pc = ws.cell(row=r, column=3, value=_pct(val))
        pc.number_format = pct_fmt
        pc.fill = fill
        pc.font = font
        row_num[0] += 1

    _hdr()

    _sec_hdr("INKOMEN", C_INC_HDR)
    for cat in INCOME_CATS:
        _data_row(cat, C_INC_ROW)
    _subtotal("Totaal Inkomen", INCOME_CATS, C_INC_SUB)
    _blank()

    _sec_hdr("VASTE LASTEN", C_VL_HDR)
    for cat in VASTE_LASTEN_CATS:
        _data_row(cat, C_VL_ROW)
    _subtotal("Totaal Vaste Lasten", VASTE_LASTEN_CATS, C_VL_SUB)
    _blank()

    _sec_hdr("DAGELIJKSE UITGAVEN", C_DAG_HDR)
    for cat in DAGELIJKS_CATS:
        _data_row(cat, C_DAG_ROW)
    _subtotal("Totaal Dagelijks", DAGELIJKS_CATS, C_DAG_SUB)
    _blank()

    _sec_hdr("OVERIG", C_OVR_HDR)
    for cat in OVERIG_CATS:
        _data_row(cat, C_OVR_ROW)
    _subtotal("Totaal Overig", OVERIG_CATS, C_OVR_SUB)
    _blank()
    _blank()

    _sec_hdr("SAMENVATTING", C_SAM_HDR)
    _sam_row("Totaal Inkomen",      INCOME_CATS,       C_INC_SUB)
    _sam_row("Totaal Vaste Lasten", VASTE_LASTEN_CATS, C_VL_SUB)
    _sam_row("Totaal Dagelijks",    DAGELIJKS_CATS,    C_DAG_SUB)
    _sam_row("Totaal Overig",       OVERIG_CATS,       C_OVR_SUB)

    r     = row_num[0]
    netto = df_real["amount"].sum()
    color = C_NETTO_POS if netto >= 0 else C_NETTO_NEG
    fill  = _fill(color)
    font  = Font(bold=True, color=WHITE, name="Arial", size=11)
    ws.cell(row=r, column=1, value="NETTO").fill = fill
    ws.cell(row=r, column=1).font = font
    tc = ws.cell(row=r, column=2, value=round(netto))
    tc.number_format = euro_fmt
    tc.fill = fill
    tc.font = font
    pc = ws.cell(row=r, column=3, value=_pct(netto))
    pc.number_format = pct_fmt
    pc.fill = fill
    pc.font = font
    ws.row_dimensions[r].height = 18

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 16


def write_unknowns_sheet(ws, df):
    ws.title = "Onbekend"
    unknowns = df[df["category"] == "Onbekend"].copy()

    if unknowns.empty:
        ws.cell(row=1, column=1,
                value="Alles gecategoriseerd - geen onbekende transacties!")
        ws.cell(row=1, column=1).font = Font(name="Arial", size=10, bold=True,
                                              color=C_INC_HDR)
        return

    # Tip row
    tip = ("Voeg een keyword toe aan rules.xlsx en herrun process.py "
           "om deze merchants automatisch te categoriseren.")
    ws.cell(row=1, column=1, value=tip)
    ws.cell(row=1, column=1).font = Font(name="Arial", size=9, italic=True,
                                          color="595959")
    ws.merge_cells("A1:E1")
    ws.row_dimensions[1].height = 14

    headers = ["Merchant", "Aantal", "Totaal (EUR)",
               "Voorbeeld omschrijving", "Jouw categorie (vul in)"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=c, value=h)
        cell.fill = _fill("C55A11")
        cell.font = Font(bold=True, color=WHITE, name="Arial", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

    grouped = (
        unknowns.groupby("merchant", sort=False)
        .agg(aantal=("amount", "count"),
             totaal=("amount", "sum"),
             voorbeeld=("description", "first"))
        .reset_index()
        .sort_values("totaal", key=lambda x: x.abs(), ascending=False)
    )

    unk_fill = _fill(C_UNKNOWN)
    euro_fmt = '€ #,##0.00;-€ #,##0.00'

    for r, (_, row) in enumerate(grouped.iterrows(), 3):
        vals = [row["merchant"], int(row["aantal"]), row["totaal"],
                str(row["voorbeeld"])[:80], ""]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.fill = unk_fill
            cell.font = Font(name="Arial", size=9)
            if c == 3:
                cell.number_format = euro_fmt
            if c == 5:
                cell.font = Font(name="Arial", size=9, italic=True,
                                  color="595959")

    for c, w in enumerate([28, 8, 14, 55, 28], 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    last_data_row = 2 + len(grouped)
    ws.auto_filter.ref = f"A2:{get_column_letter(len(headers))}{last_data_row}"


def write_controle_sheet(ws, df, year_label=""):
    """Per-month reconciliation: categorised + onbekend + interne OB = netto."""
    ws.title = f"Controle{' ' + year_label if year_label else ''}"
    ws.freeze_panes = "B2"
    euro_fmt = '€ #,##0;-€ #,##0'

    periods = sorted(df["month"].unique())
    df_real = df[df["category"] != "Interne Overboeking"]
    df_int  = df[df["category"] == "Interne Overboeking"]

    def _s(frame, period):
        return float(frame[frame["month"] == period]["amount"].sum())

    headers = [
        "Maand", "Inkomen (cat)", "Uitgaven (cat)",
        "Onbekend Inc", "Onbekend Uit", "Netto Alle", "Interne OB",
    ]

    r = 1
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=r, column=c, value=h)
        cell.fill = _fill(C_HEADER)
        cell.font = Font(bold=True, color=WHITE, name="Arial", size=9)
        cell.alignment = Alignment(horizontal="center" if c > 1 else "left")
    ws.row_dimensions[r].height = 18

    EXPENSE_CATS = VASTE_LASTEN_CATS + DAGELIJKS_CATS + OVERIG_CATS
    running = [0.0] * (len(headers) - 1)

    for period in periods:
        r += 1
        inc_cat = _s(df_real[df_real["category"].isin(INCOME_CATS)], period)
        uit_cat = _s(df_real[df_real["category"].isin(EXPENSE_CATS)], period)
        unk_inc = _s(df_real[(df_real["category"] == "Onbekend") & (df_real["amount"] > 0)], period)
        unk_uit = _s(df_real[(df_real["category"] == "Onbekend") & (df_real["amount"] < 0)], period)
        netto   = _s(df_real, period)
        interne = _s(df_int, period)

        vals = [inc_cat, uit_cat, unk_inc, unk_uit, netto, interne]
        ws.cell(row=r, column=1, value=format_period(period, "month")).font = Font(name="Arial", size=9)
        for c, v in enumerate(vals, 2):
            cell = ws.cell(row=r, column=c, value=round(v))
            cell.number_format = euro_fmt
            cell.font = Font(name="Arial", size=9)
            cell.alignment = Alignment(horizontal="right")
        for i, v in enumerate(vals):
            running[i] += v

    r += 1
    ws.cell(row=r, column=1, value="Totaal").font = Font(bold=True, name="Arial", size=9)
    for c, v in enumerate(running, 2):
        cell = ws.cell(row=r, column=c, value=round(v))
        cell.number_format = euro_fmt
        cell.font = Font(bold=True, name="Arial", size=9)
        cell.alignment = Alignment(horizontal="right")

    ws.column_dimensions["A"].width = 12
    for c in range(2, len(headers) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 15


def save_output(df, year_label=""):
    suffix = f" {year_label}" if year_label else ""
    wb = Workbook()

    ws_tx = wb.active
    write_transactions_sheet(ws_tx, df)
    ws_tx.title = f"Transacties{suffix}"

    ws_mo = wb.create_sheet()
    write_overview_sheet(ws_mo, df, group_by="month")
    ws_mo.title = f"Maand Overzicht{suffix}"

    ws_js = wb.create_sheet()
    write_jaar_samenvatting_sheet(ws_js, df, year_label=year_label)

    ws_unk = wb.create_sheet()
    write_unknowns_sheet(ws_unk, df)
    ws_unk.title = f"Onbekend{suffix}"

    ws_ctrl = wb.create_sheet()
    write_controle_sheet(ws_ctrl, df, year_label=year_label)

    if year_label:
        filename = f"boekhouding_{year_label}.xlsx"
    else:
        now = datetime.now().strftime("%Y%m%d_%H%M")
        filename = f"boekhouding_{now}.xlsx"

    out_path = os.path.join(OUTPUT_DIR, filename)
    wb.save(out_path)
    print(f"Opgeslagen: {out_path}")
    return out_path


# ── Step 4: Create starter rules.xlsx ────────────────────────────────────────
def create_starter_rules():
    starter = [
        # ── INKOMEN ──────────────────────────────────────────────────────────
        ("salaris",                 "Salaris"),
        ("hogeschool",              "Salaris"),
        ("loon",                    "Salaris"),
        ("duo",                     "DUO / Studiefinanciering"),
        ("studiefinanciering",      "DUO / Studiefinanciering"),
        ("zorgtoeslag",             "Zorgtoeslag"),
        ("toeslagen",               "Zorgtoeslag"),
        # ── VASTE LASTEN ─────────────────────────────────────────────────────
        ("huiver",                  "Huur"),
        ("kamer",                   "Huur"),
        ("unive",                   "Zorgverzekering"),
        ("cz ",                     "Zorgverzekering"),
        ("zilveren kruis",          "Zorgverzekering"),
        ("youfone",                 "Telefoon & Internet"),
        ("t-mobile",                "Telefoon & Internet"),
        ("abn amro",                "Bankkosten"),
        ("spotify",                 "Abonnementen"),
        ("netflix",                 "Abonnementen"),
        ("ziggo",                   "Abonnementen"),
        ("sportcity",               "Sport & Fitness"),
        ("david lloyd",             "Sport & Fitness"),
        # ── DAGELIJKS ────────────────────────────────────────────────────────
        ("albert heijn",            "Boodschappen"),
        ("dirk",                    "Boodschappen"),
        ("jumbo",                   "Boodschappen"),
        ("lidl",                    "Boodschappen"),
        ("ah ",                     "Boodschappen"),
        ("tuda fruta",              "Boodschappen"),
        ("takeaway",                "Eten & Drinken"),
        ("deliveroo",               "Eten & Drinken"),
        ("ming kee",                "Eten & Drinken"),
        ("cn bagijn",               "Eten & Drinken"),
        ("koffiehuis",              "Eten & Drinken"),
        ("crow-bar",                "Eten & Drinken"),
        ("ls nine bar",             "Eten & Drinken"),
        ("cafe",                    "Eten & Drinken"),
        ("tabakshop",               "Eten & Drinken"),
        ("gogo tabak",              "Eten & Drinken"),
        ("ns groep",                "OV & Reizen"),
        ("ns reizigers",            "OV & Reizen"),
        ("bck*ns",                  "OV & Reizen"),
        ("den haag cs",             "OV & Reizen"),
        ("bol.com",                 "Dagelijks Overig"),
        ("media markt",             "Dagelijks Overig"),
        ("heilzaam",                "Gezondheid"),
        ("apotheek",                "Gezondheid"),
        ("laurenskerk",             "Cultuur & Entertainment"),
        # ── OVERIG ───────────────────────────────────────────────────────────
        ("sparen",                  "Sparen"),
        ("belastingdienst",         "Diversen"),
        ("primera",                 "Diversen"),
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "Rules"
    ws.cell(row=1, column=1, value="keyword")
    ws.cell(row=1, column=2, value="category")
    for c in [1, 2]:
        ws.cell(row=1, column=c).fill = _fill(C_HEADER)
        ws.cell(row=1, column=c).font = Font(bold=True, color=WHITE, name="Arial")

    for r, (kw, cat) in enumerate(starter, 2):
        ws.cell(row=r, column=1, value=kw)
        ws.cell(row=r, column=2, value=cat)

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 30

    ws2 = wb.create_sheet("Instructies")
    instructions = [
        ("Hoe werkt rules.xlsx?", ""),
        ("", ""),
        ("keyword",  "De tekst in de merchant naam (niet hoofdlettergevoelig, deelwoord)"),
        ("category", "De post die je aan die transactie wilt geven"),
        ("", ""),
        ("Tips:", ""),
        ("Volgorde telt",      "De EERSTE match wint — zet specifieke regels bovenaan"),
        ("Inkomsten",          "Gebruik 'Salaris', 'DUO / Studiefinanciering' etc."),
        ("Nieuw toevoegen",    "Voeg een rij toe aan het Rules tabblad en herrun"),
    ]
    for r, (a, b) in enumerate(instructions, 1):
        ws2.cell(row=r, column=1, value=a).font = Font(bold=(r == 1), name="Arial")
        ws2.cell(row=r, column=2, value=b).font = Font(name="Arial")
    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 65
    wb.save(RULES_FILE)
    print(f"Starter rules.xlsx aangemaakt: {RULES_FILE}")


# ── Step 5: Migrate existing rules.xlsx to new category names ────────────────
CATEGORY_MIGRATION = {
    "inkomen":                   "Salaris",
    "inkomsten":                 "Salaris",
    "inkomsten overig":          "Inkomsten Overig",
    "verzekeringen":             "Zorgverzekering",
    "telefoon":                  "Telefoon & Internet",
    "kamerhuur":                 "Huur",
    "huur & wonen":              "Huur",
    "vaste lasten":              "Abonnementen",
    "belastingen":               "Diversen",
    "sport":                     "Sport & Fitness",
    "ov & reizen":               "OV & Reizen",
    "gezondheid":                "Gezondheid",
    "boodschappen":              "Boodschappen",
    "eten & drinken":            "Eten & Drinken",
    "bankkosten":                "Bankkosten",
    "diversen":                  "Diversen",
    "sparen":                    "Sparen",
    "online winkelen":           "Dagelijks Overig",
    "tabak":                     "Eten & Drinken",
    "cultuur & entertainment":   "Cultuur & Entertainment",
}

KEYWORD_OVERRIDES = {
    "tabakshop":         "Eten & Drinken",
    "gogo tabak":        "Eten & Drinken",
    "bol.com":           "Dagelijks Overig",
    "media markt":       "Dagelijks Overig",
    "laurenskerk":       "Cultuur & Entertainment",
    "belastingdienst":   "Diversen",
    "sportcity":         "Sport & Fitness",
    "david lloyd":       "Sport & Fitness",
    "youfone":           "Telefoon & Internet",
    "unive":             "Zorgverzekering",
    "huiver":            "Huur",
    "hogeschool":        "Salaris",
    "salaris":           "Salaris",
    "abn amro":          "Bankkosten",
}


def migrate_rules():
    """Update existing rules.xlsx to the new category taxonomy."""
    if not os.path.exists(RULES_FILE):
        print("Geen rules.xlsx gevonden.")
        return

    wb = load_workbook(RULES_FILE)
    if "Rules" not in wb.sheetnames:
        print("Geen 'Rules' tabblad gevonden in rules.xlsx.")
        return

    ws      = wb["Rules"]
    changed = 0
    for row in ws.iter_rows(min_row=2):
        kw_cell  = row[0]
        cat_cell = row[1]
        kw  = str(kw_cell.value  or "").strip().lower()
        old = str(cat_cell.value or "").strip()

        if kw in KEYWORD_OVERRIDES:
            new = KEYWORD_OVERRIDES[kw]
        else:
            new = CATEGORY_MIGRATION.get(old.lower(), old)

        if new != old:
            cat_cell.value = new
            changed += 1

    wb.save(RULES_FILE)
    print(f"rules.xlsx bijgewerkt: {changed} categorie(en) hernoemd")


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    if "--migrate-rules" in sys.argv:
        migrate_rules()
        return

    if "--create-rules" in sys.argv or not os.path.exists(RULES_FILE):
        print("Starter rules.xlsx aanmaken ...")
        create_starter_rules()
        if "--create-rules" in sys.argv:
            return

    # ── Parse --year flag ─────────────────────────────────────────────────────
    year_filter = None
    if "--year" in sys.argv:
        idx = sys.argv.index("--year")
        if idx + 1 < len(sys.argv):
            year_filter = sys.argv[idx + 1]

    print("\nABN AMRO Bookkeeping Processor")
    print("=" * 38)
    if year_filter:
        print(f"Jaar filter: {year_filter}")

    files = find_input_files()
    df    = load_transactions(files)

    if year_filter:
        df = df[df["year"] == year_filter].reset_index(drop=True)
        if df.empty:
            sys.exit(f"Geen transacties gevonden voor jaar {year_filter}")
        print(f"Gefilterd op {year_filter}: {len(df)} transacties")

    rules = load_rules()
    df    = apply_categories(df, rules)
    df    = detect_internal_transfers(df)
    out   = save_output(df, year_label=year_filter or "")

    # ── Terminal summary ──────────────────────────────────────────────────────
    df_real = df[df["category"] != "Interne Overboeking"]
    print()
    for m in sorted(df_real["month"].unique()):
        mdf  = df_real[df_real["month"] == m]
        inc  = mdf[mdf["category"].isin(INCOME_CATS)]["amount"].sum()
        vl   = mdf[mdf["category"].isin(VASTE_LASTEN_CATS)]["amount"].sum()
        dag  = mdf[mdf["category"].isin(DAGELIJKS_CATS)]["amount"].sum()
        net  = mdf["amount"].sum()
        yr, mo = m.split("-")
        label  = f"{MAANDEN_NL[int(mo)]} {yr}"
        print(f"  {label:<10}  |  in: {dutch_euros(inc):<10}"
              f"  |  vaste lasten: {dutch_euros(vl):<10}"
              f"  |  dagelijks: {dutch_euros(dag):<10}"
              f"  |  netto: {dutch_euros(net)}")

    t_inc = df_real[df_real["category"].isin(INCOME_CATS)]["amount"].sum()
    t_vl  = df_real[df_real["category"].isin(VASTE_LASTEN_CATS)]["amount"].sum()
    t_dag = df_real[df_real["category"].isin(DAGELIJKS_CATS)]["amount"].sum()
    t_net = df_real["amount"].sum()
    print("  " + "-" * 80)
    print(f"  {'Totaal':<10}  |  in: {dutch_euros(t_inc):<10}"
          f"  |  vaste lasten: {dutch_euros(t_vl):<10}"
          f"  |  dagelijks: {dutch_euros(t_dag):<10}"
          f"  |  netto: {dutch_euros(t_net)}")
    print(f"\n  Output: {out}\n")


if __name__ == "__main__":
    main()
